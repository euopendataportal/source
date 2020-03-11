#    Copyright (C) <2018>  <Publications Office of the European Union>
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU Affero General Public License as
#    published by the Free Software Foundation, either version 3 of the
#    License, or (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU Affero General Public License for more details.
#
#    You should have received a copy of the GNU Affero General Public License
#    along with this program.  If not, see <https://www.gnu.org/licenses/>.
#
#    contact: <https://publications.europa.eu/en/web/about-us/contact>

import cPickle as pickle
import copy
import datetime
import logging
import shutil
import time
import traceback
import ujson as json
import uuid
import zipfile
from io import BytesIO
from urllib import urlencode

import ckan.controllers.package as core_package
import ckan.lib.base as base
import ckan.lib.datapreview as datapreview
import ckan.lib.helpers as core_helpers
import ckan.lib.maintain as maintain
import ckan.lib.navl.dictization_functions
import ckan.lib.navl.dictization_functions as dict_func
import ckanext.ecportal.lib.search.dcat_index as index
import ckan.logic as logic
import ckan.model as model
import ckan.plugins as p
import ckan.plugins.toolkit as tk
import os
import pylons.controllers.util as util
import re
import sqlalchemy
from ckan import plugins
from ckan.common import OrderedDict, _, request, c, g, response
from ckan.controllers.package import PackageController
from ckan.lib.search import SearchIndexError
from genshi.template import MarkupTemplate
from openpyxl import Workbook
from openpyxl import load_workbook
from os.path import basename
from paste.deploy.converters import asbool
from pylons import config, session
from doi.facade import doi_facade
import ckanext.ecportal.configuration.configuration_constants as constants

import ckanext.ecportal.action.customsearch as customsearch
import ckanext.ecportal.action.ecportal_get as ecportal_action
import ckanext.ecportal.action.ecportal_validation as validation
import ckanext.ecportal.helpers as ckanext_helpers
import ckanext.ecportal.lib.converters as ckanext_converters
import ckanext.ecportal.lib.ingestion.ingestion_package as ingestion
import ckanext.ecportal.lib.page_util as page_util
import ckanext.ecportal.lib.search as search
import ckanext.ecportal.lib.ui_util as ui_util
import ckanext.ecportal.lib.uri_util as uri_util
from ckanext.ecportal.action import ecportal_save
from ckanext.ecportal.action.ecportal_save import package_save
from ckanext.ecportal.forms import EXCEL_FORMAT
from ckanext.ecportal.forms import FORMATS
from ckanext.ecportal.forms import JSON_FORMAT
from ckanext.ecportal.forms import RDF_FORMAT
from ckanext.ecportal.lib.cache import redis_cache
from ckanext.ecportal.lib.controlled_vocabulary_util import Distribution_controlled_vocabulary, \
    Documentation_controlled_vocabulary, retrieve_all_distribution_types, retrieve_all_documentation_types
from ckanext.ecportal.model import iso639_1_bcp47converison

from ckanext.ecportal.model.catalog_dcatapop import CatalogDcatApOp
from ckanext.ecportal.migration.migration_constants import DATASET_URI_PREFIX
from ckanext.ecportal.model.dataset_dcatapop import DatasetDcatApOp, SchemaGeneric, ResourceValue, DCATAPOP_PRIVATE_GRAPH_NAME, \
    DCATAPOP_PUBLIC_GRAPH_NAME, DCATAPOP_PRIVATE_DATASET, DCATAPOP_PUBLIC_DATASET
from ckanext.ecportal.model.ecodp_package_contact_info import Package_contact_info
from ckanext.ecportal.virtuoso import PRIVACY_STATE_PRIVATE, PRIVACY_STATE_PUBLIC
from ckanext.ecportal.configuration.configuration_constants import CKAN_PATH

FIX_PATH = '/pairtree_root/de/fa/ul/t/obj/'

check_access = logic.check_access
render = base.render
abort = util.abort
redirect = base.redirect
NotFound = logic.NotFound
NotAuthorized = logic.NotAuthorized
ValidationError = logic.ValidationError
get_action = logic.get_action
lookup_package_plugin = ckan.lib.plugins.lookup_package_plugin
indexer = index.PackageSearchIndex()

log = logging.getLogger(__name__)
_validate = ckan.lib.navl.dictization_functions.validate
_check_access = logic.check_access
_and_ = sqlalchemy.and_

QUERY_FIELDS = "name^4 title^4 tags^2 groups^2 text"

export_excel_order = []

export_excel_lang_column = {}

locales = ckanext_helpers.get_available_locales()

excel_multi_values_separator = ' ,'


def search_url(params, package_type=None):
    if not package_type or package_type == 'dataset':
        url = core_helpers.url_for(
            controller='ckanext.ecportal.controllers.package:ECPORTALPackageController',
            action='search')
    else:
        url = core_helpers.url_for('{0}_search'.format(package_type))
    return core_package.url_with_params(url, params)


def retrieve_form_schema_list_for_datasets(datasets_list):
    form_schema_list = []
    for dataset in datasets_list:
        form_schema = ui_util.transform_dcat_schema_to_form_schema(dataset)
        form_schema_list.append(form_schema)
    return form_schema_list


def retrieve_bulk_edit_form_schema(datasets_keys_diff, form_schema_list):
    bulk_filtered_form_schema = {}
    form_schema_reference = form_schema_list[0]
    for key, value in form_schema_reference.iteritems():
        is_property_different = key in datasets_keys_diff
        if is_property_different:
            if isinstance(value, list):
                bulk_filtered_form_schema[key] = []
            elif isinstance(value, dict):
                bulk_filtered_form_schema[key] = {}
            else:
                bulk_filtered_form_schema[key] = ""
        else:
            bulk_filtered_form_schema[key] = value
    return bulk_filtered_form_schema


def retrieve_differences_for_form_schema_list(form_schema_list):
    # This method is used to compare the current model of form schema which is compose of nested structure
    # containing lists and dicts
    datasets_keys_diff = []

    for form_schema in form_schema_list:
        variables_existence_diff = set(form_schema_list[0].keys()) - set(form_schema.keys())
        datasets_keys_diff = datasets_keys_diff + (list(variables_existence_diff))

        for key, value in form_schema.iteritems():
            try:
                reference_value = form_schema_list[0].get(key)
                if reference_value is None or not isinstance(reference_value, type(value)):
                    datasets_keys_diff.append(key)
                else:
                    if isinstance(value, dict):
                        value = value.iteritems()
                        if not isinstance(reference_value, dict):
                            log.error("Model is corrupted, value {0} must be a dict".format(reference_value))
                            raise Exception("Model is corrupted, property should be a dict")
                        reference_value = reference_value.iteritems()
                    if isinstance(value, list):
                        if len(value) > 0 and isinstance(value[0], dict):
                            diff = retrieve_differences_for_form_schema_list(value + reference_value)
                            if diff:
                                datasets_keys_diff.append(key)
                            continue
                    if set(reference_value) - set(value):
                        datasets_keys_diff.append(key)
            except Exception as e:
                log.warn("Error retrieving differences between datasets: {0} at key: {1}, value: {2}", e, key, value)
    return datasets_keys_diff


def is_private(request):
    privacy_parameter = request.POST.get('save')
    return privacy_parameter == u'Save as draft'


def patch_and_save_datasets(context, data_dict, datasets_list, request):
    '''

    :param context:
    :param dict data_dict: dict with key/value of changes to be applyed to each dataset
    :param datasets_list: list of datasets in scope of this change
    :param request: the original post data
    :return:
    '''

    is_private_state = is_private(request)
    if is_private_state:
        data_dict['private'] = 'True'
    for dataset in datasets_list:  # type: DatasetDcatApOp
        old_dataset = pickle.dumps(dataset)
        dataset.patch_dataset_for_package_dict(data_dict, {}, context)
        ecportal_save.update_exisiting_dataset(dataset, pickle.loads(old_dataset), context)


class ECPORTALPackageController(PackageController):
    '''
    Custom dataset search for UI
    '''
    plugins.implements(plugins.IDatasetForm)

    RESOURCES_TYPES = [
        [Documentation_controlled_vocabulary.DOCUMENTATION_MAIN, _('Documentation: Main')],
        [Documentation_controlled_vocabulary.DOCUMENTATION_RELATED,
         _('Documentation: Related')],
        [Documentation_controlled_vocabulary.WEBPAGE_RELATED, _('Documentation: Webpage')],
        [Distribution_controlled_vocabulary.FEED_INFO, _('Distribution: Feed')],
        [Distribution_controlled_vocabulary.WEB_SERVICE, _('Distribution: Web Service')],
        [Distribution_controlled_vocabulary.DOWNLOADABLE_FILE, _('Distribution: Download')],
        [Distribution_controlled_vocabulary.VISUALIZATION, _('Visualization')]
    ]

    RESOURCES_TYPES_DOCUMENTATION = [
        [Documentation_controlled_vocabulary.DOCUMENTATION_MAIN, _('Documentation: Main')],
        [Documentation_controlled_vocabulary.DOCUMENTATION_RELATED,
         _('Documentation: Related')],
        [Documentation_controlled_vocabulary.WEBPAGE_RELATED, _('Documentation: Webpage')]
    ]

    RESOURCES_TYPES_DISTRIBUTION = [
        [Distribution_controlled_vocabulary.FEED_INFO, _('Distribution: Feed')],
        [Distribution_controlled_vocabulary.WEB_SERVICE, _('Distribution: Web Service')],
        [Distribution_controlled_vocabulary.DOWNLOADABLE_FILE, _('Distribution: Download')]
    ]

    RESOURCES_TYPES_VISUALIZATION = [
        [Distribution_controlled_vocabulary.VISUALIZATION, _('Visualization')]
    ]

    def resource_read(self, id, resource_id):
        context = {'model': model, 'session': model.Session,
                   'user': c.user or c.author, 'auth_user_obj': c.userobj, 'for_view': True}
        language = tk.request.environ['CKAN_LANG'] or config.get('ckan.locale_default', 'en')

        # dataset_uri = tsch.get_uri_based_on_property_value("dcatapop-public","<http://data.europa.eu/88u/ontology/dcatapop#ckanName>", id)
        # if dataset_uri:
        datas = {}
        try:
            datas = ecportal_action.package_show(context, {"id": id})
        except NotFound as e:
            import traceback
            log.error('{0}'.format(e))
            log.error(traceback.print_exc())
            base.abort(404, "package does not exists")

        rs = datas['resources']
        c.resource = next((r for r in rs if r['id'] == resource_id), None)
        if not c.resource:
            base.abort(404, "resource does not exists")

        c.package = context['package']
        c.pkg = c.pkg_dict
        c.pkg_dict = datas

        try:
            prev_url = next((download for download in c.resource.get('download_url', [])), {}) or next(
                (download for download in c.resource.get('access_url', [])), {})
            prev_format = c.resource.get('format', '')
            prev_res = {'url': prev_url,
                        'format': prev_format,
                        'id': resource_id}
            c.prev_res = prev_res
            c.resource['can_be_previewed'] = self._resource_preview({'resource': prev_res, 'package': datas})

            # Generate json-ld for dataset
            json_ld = ui_util.load_json_ld_resource(c.resource)
            c.resource['json_ld'] = json_ld

            return render('package/resource_read.html')

        # else:
        #     base.abort(404, "package does not exists")
        except Exception as e:
            import traceback
            log.error('{0}'.format(e))
            log.error(traceback.print_exc())

    def _search_template(self, package_type):
        return super(ECPORTALPackageController, self)._search_template(package_type)

    def _resource_preview(self, data_dict):
        return bool(datapreview.res_format(
            data_dict['resource']) in datapreview.direct() + datapreview.loadable() or datapreview.get_preview_plugin(
            data_dict, return_first=True))

    def read(self, id, format='html'):
        ctype, extension, loader = self._content_type_from_extension(format)
        if not format == 'html':
            if not ctype:
                # An unknown format, we'll carry on in case it is a
                # revision specifier and re-constitute the original id
                id = "%s.%s" % (id, format)
                ctype, format, loader = "text/html; charset=utf-8", "html", \
                                        MarkupTemplate
        else:
            ctype, format, loader = self._content_type_from_accept()

        response.headers['Content-Type'] = ctype

        package_type = self._get_package_type(id.split('@')[0])
        context = {'user': c.user or c.author, 'for_view': extension != 'rdf',
                   'auth_user_obj': c.userobj}
        data_dict = {'id': id}

        # interpret @<revision_id> or @<date> suffix
        split = id.split('@')
        if len(split) == 2:
            data_dict['id'], revision_ref = split
            if model.is_id(revision_ref):
                context['revision_id'] = revision_ref
            else:
                try:
                    date = core_helpers.date_str_to_datetime(revision_ref)
                    context['revision_date'] = date
                except TypeError, e:
                    abort(400, _('Invalid revision format: %r') % e.args)
                except ValueError, e:
                    abort(400, _('Invalid revision format: %r') % e.args)
        elif len(split) > 2:
            abort(400, _('Invalid revision format: %r') %
                  'Too many "@" symbols')

        # check if package exists
        try:
            language = tk.request.environ['CKAN_LANG'] or config.get('ckan.locale_default', 'en')
            c.pkg_dict = get_action('package_show')(context, data_dict)

            c.pkg_dict['citation_styles'] = constants.DOI_CONFIG.citation_formats
            doi = c.pkg_dict.get('doi', '')
            if doi:
                c.pkg_dict['citation'] = ui_util.get_citation(doi, iso639_1_bcp47converison.ISO_639_1_TO_BCP47[language])

            c.package = context['package']  # type DatasetDcatApOp
            c.pkg = c.pkg_dict
        except NotFound:
            abort(404, _('Dataset not found'))
        except NotAuthorized:
            abort(401, _('Unauthorized to read package %s') % id)
        except Exception as e:
            import traceback
            log.error(traceback.print_exc())
            abort(400, _('Data Error. Could not retrieve and transform dataset'))

        # TODO can the resources be previewed?
        # for resource in c.pkg_dict['resources']:
        #     resource['can_be_previewed'] = self._resource_preview(
        #         {'resource': resource, 'package': c.pkg_dict})

        # Generate json-ld for dataset
        json_ld = ui_util.load_json_ld_dataset(c.pkg_dict)
        c.pkg['json_ld'] = json_ld
        self._setup_template_variables(context, {'id': id},
                                       package_type=package_type)

        # package_saver.PackageSaver().render_package(c.pkg_dict, context)

        template = self._read_template(package_type)
        template = template[:template.index('.') + 1] + format

        if 'rdf' == format:
            return c.package.get_dataset_as_rdfxml()

        try:
            return render(template, loader_class=loader)
        except ckan.lib.render.TemplateNotFound:
            msg = _("Viewing {package_type} datasets in {format} format is "
                    "not supported (template file {file} not found).".format(
                package_type=package_type, format=format, file=template))
            abort(404, msg)

        assert False, "We should never get here"

    def search(self):
        from ckan.lib.search import SearchError

        package_type = self._guess_package_type()
        before_start_time = time.time()
        try:
            context = {'model': model, 'user': c.user or c.author,
                       'auth_user_obj': c.userobj}
            check_access('site_read', context)
        except NotAuthorized:
            abort(401, _('Not authorized to see this page'))

        timing = time.time()
        if request.GET.get('ext_boolean') in ['all', 'any', 'exact']:
            session['ext_boolean'] = request.GET['ext_boolean']
            session.save()

        # unicode format (decoded from utf8)
        q = c.q = request.params.get('q', u'')
        c.query_error = False
        try:
            page = int(request.params.get('page', 1))
        except ValueError, e:
            abort(400, ('"page" parameter must be an integer'))
        limit = g.datasets_per_page

        # most search operations should reset the page counter:
        params_nopage = [(k, v) for k, v in request.params.items()
                         if k != 'page']

        new_params_nopage = [];
        for key, value in params_nopage:
            if key == 'eurovoc_domains':
                new_params_nopage.append(('eurovoc_domains', value))
            else:
                new_params_nopage.append((key, value))

        params_nopage = new_params_nopage;

        def drill_down_url(alternative_url=None, **by):
            return core_helpers.add_url_param(alternative_url=alternative_url,
                                              controller='package', action='search',
                                              new_params=by)

        c.drill_down_url = drill_down_url

        def remove_field(key, value=None, replace=None):
            return core_helpers.remove_url_param(key, value=value, replace=replace,
                                                 controller='package', action='search')

        c.remove_field = remove_field

        sort_by = request.params.get('sort', None)
        params_nosort = [(k, v) for k, v in params_nopage if k != 'sort']

        def _sort_by(fields):
            """
            Sort by the given list of fields.

            Each entry in the list is a 2-tuple: (fieldname, sort_order)

            eg - [('metadata_modified', 'desc'), ('name', 'asc')]

            If fields is empty, then the default ordering is used.
            """
            params = params_nosort[:]

            if fields:
                sort_string = ', '.join('%s %s' % f for f in fields)
                params.append(('sort', sort_string))
            return search_url(params, package_type)

        c.sort_by = _sort_by
        if not sort_by:
            c.sort_by_fields = []
        else:
            c.sort_by_fields = [field.split()[0]
                                for field in sort_by.split(',')]

        def pager_url(q=None, page=None):
            params = list(params_nopage)
            params.append(('page', page))
            return core_package.search_url(params, package_type)

        c.search_url_params = urlencode(core_package._encode_params(params_nopage))

        try:
            c.fields = []
            # c.fields_grouped will contain a dict of params containing
            # a list of values eg {'tags':['tag1', 'tag2']}
            c.fields_grouped = {}
            search_extras = {}
            fq = ''
            for (param, value) in request.params.items():
                if param not in ['q', 'page', 'sort'] \
                        and len(value) and not param.startswith('_'):
                    if not param.startswith('ext_'):
                        c.fields.append((param, value))
                        paramFQ = 'eurovoc_domains' if (param == 'eurovoc_domains') else param;
                        fq += ' %s:"%s"' % (paramFQ, value)
                        if param not in c.fields_grouped:
                            c.fields_grouped[param] = [value]
                        else:
                            c.fields_grouped[param].append(value)
                    else:
                        search_extras[param] = value

            context = {'model': model, 'session': model.Session,
                       'user': c.user or c.author, 'for_view': True,
                       'auth_user_obj': c.userobj}

            if package_type and package_type != 'dataset':
                # Only show datasets of this particular type
                fq += ' +dataset_type:{type}'.format(type=package_type)
            else:
                # Unless changed via config options, don't show non standard
                # dataset types on the default search page
                if not asbool(config.get('ckan.search.show_all_types', 'False')):
                    fq += ' +dataset_type:dataset'

            facets = OrderedDict()

            default_facet_titles = {
                'organization': _('Organizations'),
                'groups': _('Groups'),
                'tags': _('Tags'),
                'res_format': _('Formats'),
                'license_id': _('Licenses'),
            }

            for facet in g.facets:
                if facet in default_facet_titles:
                    facets[facet] = default_facet_titles[facet]
                else:
                    facets[facet] = facet

            # Facet titles
            for plugin in p.PluginImplementations(p.IFacets):
                facets = plugin.dataset_facets(facets, package_type)

            c.facet_titles = facets

            data_dict = {
                'q': q,
                'fq': fq.strip(),
                'facet.field': facets.keys(),
                'rows': limit,
                'start': (page - 1) * limit,
                'sort': sort_by,
                'extras': search_extras
            }

            result_list = []
            if sort_by == 'modified_date desc':
                # This is the customized part for ODP-570
                # add the group parameter to the solr query
                data_dict.pop('start')
                data_dict['group'] = 'true'
                data_dict['group.query'] = [
                    '-organization:estat AND -organization:comp AND -organization:grow', 'organization:estat',
                    'organization:comp', 'organization:grow']
                data_dict['group.format'] = 'simple'
                data_dict['rows'] = 2147483646

                before_duration = time.time() - before_start_time
                log.info("Controller prepare modified_date search took {0}".format(before_duration))

                start_time = time.time()
                query = get_action('custom_package_search')(context, data_dict)
                duration = time.time() - start_time
                log.info("Call custom_package_search action took {0}".format(duration))
                cached_result = []

                for name, group in query['groups'].iteritems():
                    cached_result += group['doclist']['docs']

                start = (page - 1) * limit


                post_start_time = time.time()
                result_list = customsearch.check_solr_result(context, cached_result[start:], limit)
                post_duration = time.time() - post_start_time
                log.info("Call package_search action took {0}".format(post_duration))

            else:
                before_duration = time.time() - before_start_time
                log.info("Controller prepare other search took {0}".format(before_duration))
                start_time = time.time()
                query = get_action('package_search')(context, data_dict)
                duration = time.time() - start_time
                log.info("Call package_search action took {0}".format(duration))
                result_list = query['results']

            c.sort_by_selected = query['sort']

            c.page = page_util.Page(
                collection=result_list,
                page=page,
                url=pager_url,
                item_count=query['count'],
                items_per_page=limit
            )
            c.facets = query['facets']
            c.search_facets = query['search_facets']
            c.page.items = result_list
        except SearchError, se:
            log.error('Dataset search error: %r', se.args)
            c.query_error = True
            c.facets = {}
            c.search_facets = {}
            c.page = core_helpers.Page(collection=[])
        c.search_facets_limits = {}
        for facet in c.search_facets.keys():
            try:
                limit = int(request.params.get('_%s_limit' % facet,
                                               g.facets_default_number))
            except ValueError:
                abort(400, _('Parameter "{parameter_name}" is not '
                             'an integer').format(
                    parameter_name='_%s_limit' % facet
                ))
            c.search_facets_limits[facet] = limit

        maintain.deprecate_context_item(
            'facets',
            'Use `c.search_facets` instead.')

        self._setup_template_variables(context, {},
                                       package_type=package_type)

        log.info('Overall controller search duration before render: {0}s'.format(time.time() - timing))

        return render(self._search_template(package_type))

    def history(self, id):
        package_type = self._get_package_type(id.split('@')[0])

        if 'diff' in request.params or 'selected1' in request.params:
            try:
                params = {'id': request.params.getone('pkg_name'),
                          'diff': request.params.getone('selected1'),
                          'oldid': request.params.getone('selected2'),
                          }
            except KeyError, e:
                if 'pkg_name' in dict(request.params):
                    id = request.params.getone('pkg_name')
                c.error = \
                    _('Select two revisions before doing the comparison.')
            else:
                params['diff_entity'] = 'package'
                h.redirect_to(controller='revision', action='diff', **params)

        context = {'model': model, 'session': model.Session,
                   'user': c.user or c.author, 'auth_user_obj': c.userobj}
        data_dict = {'id': id}
        try:
            c.pkg_dict = get_action('package_show')(context, data_dict)
            c.pkg_revisions = get_action('package_revision_list')(context,
                                                                  data_dict)
            # TODO: remove
            # Still necessary for the authz check in group/layout.html
            c.pkg = context['package']

        except NotAuthorized:
            abort(401, _('Unauthorized to read package %s') % '')
        except NotFound:
            abort(404, _('Dataset not found'))

        format = request.params.get('format', '')
        if format == 'atom':
            # Generate and return Atom 1.0 document.
            from webhelpers.feedgenerator import Atom1Feed
            feed = Atom1Feed(
                title=_(u'CKAN Dataset Revision History'),
                link=h.url_for(controller='revision', action='read',
                               id=c.pkg_dict['name']),
                description=_(u'Recent changes to CKAN Dataset: ') +
                            (c.pkg_dict['title'] or ''),
                language=unicode(i18n.get_lang()),
            )
            for revision_dict in c.pkg_revisions:
                revision_date = h.date_str_to_datetime(
                    revision_dict['timestamp'])
                try:
                    dayHorizon = int(request.params.get('days'))
                except:
                    dayHorizon = 30
                dayAge = (datetime.datetime.now() - revision_date).days
                if dayAge >= dayHorizon:
                    break
                if revision_dict['message']:
                    item_title = u'%s' % revision_dict['message']. \
                        split('\n')[0]
                else:
                    item_title = u'%s' % revision_dict['id']
                item_link = h.url_for(controller='revision', action='read',
                                      id=revision_dict['id'])
                item_description = _('Log message: ')
                item_description += '%s' % (revision_dict['message'] or '')
                item_author_name = revision_dict['author']
                item_pubdate = revision_date
                feed.add_item(
                    title=item_title,
                    link=item_link,
                    description=item_description,
                    author_name=item_author_name,
                    pubdate=item_pubdate,
                )
            feed.content_type = 'application/atom+xml'
            return feed.writeString('utf-8')

        c.related_count = c.pkg.related_count
        return render(self._history_template(c.pkg_dict.get('type',
                                                            package_type)))

    def new(self, data=None, errors=None, error_summary=None):
        package_type = self.__get_package_type()

        context = {'model': 'DCATAP', 'session': model.Session,
                   'user': c.user or c.author, 'auth_user_obj': c.userobj,
                   'save': 'save' in request.params,
                   'validate': 'validate' in request.params
                   }
        if 'duplicate' in request.params:
            context['duplicate'] = request.params['duplicate']

        # Package needs to have a organization group in the call to check_access and also to save it
        try:
            logic.check_access('package_create', context)
        except logic.NotAuthorized:
            abort(401, _('Unauthorized to create a package'))

        # Check if we just want to lead the page or if we want to issue a save request
        if context['save'] and not data:
            request.POST.pop('save')
            return self._save_create_and_show_dataset_in_dashboard(context)

        data = data or self.__transform_to_data_dict(request.POST)

        if context['validate']:
            request.POST.pop('validate')
            data = data or self.__transform_to_data_dict(request.POST)
            if not data.get('title'):
                dataset = DatasetDcatApOp('')
            else:
                uri, ds_name = uri_util.new_dataset_uri_from_title(data.get('title'))
                dataset = DatasetDcatApOp(uri)
                data['name'] = ds_name
                dataset = DatasetDcatApOp(uri)
            context['package'] = dataset
            start = time.time()
            dataset.create_dataset_schema_for_package_dict(data, {}, context)
            log.info('create from dataset input took {0}s'.format((time.time() - start)))
            start = time.time()
            data = ui_util.transform_dcat_schema_to_form_schema(dataset)
            log.info('transform dataset to UI took {0}s'.format((time.time() - start)))
            dataset, errors = validation.validate_dacat_dataset(dataset, context)

        elif 'duplicate' in context and not data:  # 'and not data' really needed?
            try:
                # get the dataset
                old_data_dict = logic.get_action('package_show')(context, {'id': context['duplicate']})
                selected_dataset = context['package']  # type: DatasetDcatApOp
                dataset = DatasetDcatApOp(selected_dataset.dataset_uri)
                new_id = str(uuid.uuid4())
                # copy the dataset
                copy_dataset_dump = pickle.dumps(selected_dataset)
                copy_dataset = pickle.loads(copy_dataset_dump)  # type: DatasetDcatApOp
                copy_dataset.dataset_uri = u""
                copy_dataset.schema.uri = copy_dataset.dataset_uri
                # copy_dataset.schema.ckanName_dcatapop.get('0').value_or_uri += new_id
                copy_dataset.schema.ckanName_dcatapop = {}

                title = copy_dataset.schema.title_dcterms.get('0').value_or_uri
                for title in copy_dataset.schema.title_dcterms.values():
                    if title.lang == "en":
                        title.value_or_uri += " Duplicated"
                for distribution in copy_dataset.schema.distribution_dcat.values():
                    distribution.uri += new_id
                # copy_dataset.save_to_ts()
                # search.rebuild(copy_dataset.dataset_uri.split('/')[-1])

                context['package'] = copy_dataset
                errors = {}
                ui_pkg = ui_util.transform_dcat_schema_to_form_schema(copy_dataset)
                return self.new(ui_pkg, errors)

            except logic.NotAuthorized:
                abort(401, _('Unauthorized to read package %s') % '')
            except logic.NotFound:
                abort(404, _('Dataset not found'))

        c.resources_json = core_helpers.json.dumps(data.get('resources', []))



        catalog_list = get_action('catalogue_list')(context, {})
        language = tk.request.environ['CKAN_LANG'] or config.get('ckan.locale_default', 'en')

        catalogs = {}
        for key, value in catalog_list.items():
            title = next((title.value_or_uri for title in value.schema.title_dcterms.values() if title.lang == language), key)
            catalogs[key] = title

        c.catalogs = catalogs

        # convert tags if not supplied in data
        if data and not data.get('tag_string'):
            data['tag_string'] = ', '.join(
                core_helpers.dict_list_reduce(data.get('tags', {}), 'name'))

        errors = errors or {}
        error_summary = error_summary or {}

        # in the phased add dataset we need to know that we have already completed stage 1
        stage = ['active']
        if data.get('state') == 'draft':
            stage = ['active', 'complete']
        elif data.get('state') == 'draft-complete':
            stage = ['active', 'complete', 'complete']

        # if we are creating from a group then this allows the group to be set automatically
        data['group_id'] = request.params.get('group') or request.params.get('groups__0__id')

        # Disable the groups fields (i.e.: EuroVoc domains / Groups) in case the new dataset is created by a
        # non-admin user
        group_input_disabled = not ckanext_helpers.is_sysadmin(context['user'])

        data['resources_types'] = self.RESOURCES_TYPES;
        data['doi'] = ''

        contact_info = Package_contact_info.get_by_user(c.userobj.id or '')
        if contact_info:
            data.update(contact_info.as_dict())

        c.resources_types_documentation = list(filter(lambda x: x[0] != "http://publications.europa.eu/resource/authority/documentation-type/VISUALIZATION",
                                                      retrieve_all_documentation_types(language)))
        c.resources_types_distribution = list(filter(lambda x: x[0] != "http://publications.europa.eu/resource/authority/distribution-type/VISUALIZATION",
                                                      retrieve_all_distribution_types(language)))
        c.resources_types_visualization = self.RESOURCES_TYPES_VISUALIZATION

        vars = {'data': data, 'errors': errors,
                'error_summary': error_summary,
                'action': 'new', 'stage': stage,
                'group_input_disabled': group_input_disabled}
        errors_json = core_helpers.json.dumps(errors)
        if errors:
            log.info('Saving the dataset produced these errors:')
            log.info(errors_json)

        c.errors_json = errors_json

        self._setup_template_variables(context, {}, package_type=package_type)

        c.form = base.render(self._package_form(package_type=package_type), extra_vars=vars)

        return base.render('package/rdft_new.html')

    def pop_translation(self, dict_list):
        key_to_delete = []
        list_lang = []
        if dict_list:
            i = 0
            for res in dict_list:
                dict_lang = {}
                key_added = False
                for key, value in res.iteritems():
                    if value:
                        split_key = key.split('-')
                        if split_key[-1] in config.get('ckan.locales_offered', []):
                            lang = split_key[-1]
                            # dict_lang[split_key[-1][key[:-3]] = value
                            if lang not in dict_lang:
                                dict_lang[lang] = {}
                            dict_lang[lang][key[:-3]] = value
                            if 'iteration' not in dict_lang[lang]:
                                dict_lang[lang]['iteration'] = i
                            key_to_delete.append({'iteration': i, 'key': key})

                            key_added = True

                if key_added:
                    list_lang.append(dict_lang)
                i += 1

            for key_dict in key_to_delete:
                del dict_list[key_dict['iteration']][key_dict['key']]

        return list_lang

    def term_translation_from_pop_translation_list_lang(self, list_lang, data_dict):

        # Define reusable constants
        TITLE = 'title'
        ALT_TITLE = 'alternative_title'
        DESC = 'description'

        # TODO: refactor as soon as additional translations will be added (too much code duplications!)

        if TITLE in data_dict:
            original_title = data_dict[TITLE]

        if ALT_TITLE in data_dict:
            original_alternative_title = data_dict[ALT_TITLE]

        if DESC in data_dict:
            original_description = data_dict[DESC]

        res = []
        for dict_lang in list_lang:
            for lang_code, lang in dict_lang.iteritems():
                if original_title and lang.get(TITLE, None):
                    res.append({
                        'term': original_title,
                        'term_translation': lang[TITLE],
                        'lang_code': lang_code
                    })
                if original_alternative_title and lang.get(ALT_TITLE, None):
                    res.append({
                        'term': original_alternative_title,
                        'term_translation': lang[ALT_TITLE],
                        'lang_code': lang_code
                    })
                if original_description and lang.get(DESC, None):
                    res.append({
                        'term': original_description,
                        'term_translation': lang[DESC],
                        'lang_code': lang_code
                    })

        return res

    def _save_create_and_show_dataset_in_dashboard(self, context):
        '''
        Try to save the dataset and redirect to it when the saving is successful.
        Validation errors are presented in the page, other errors cause an abort.

        :return: Nothing, redirect to the dashboard showing the created dataset or show the occurred error.
        '''
        try:
            data_dict = self.__transform_to_data_dict(request.POST)
            old_data_dict = data_dict.copy()

            # Transform to a JSON representation and log
            log.debug("Try to save this dataset:")
            log.debug(core_helpers.dump_json(data_dict))

            errors = ""

            try:
                pkg_dict = get_action('package_create')(context, data_dict)
                if data_dict.get('doi'):
                    context['model'] = model
                    get_action('assign_doi')(context, pkg_dict)
                    if pkg_dict.privacy_state == 'public':
                        get_action('publish_doi')(context, pkg_dict)
            except ValidationError as e:
                errors = e.error_dict
                dataset = context['package']
                ui_pkg = ui_util.transform_dcat_schema_to_form_schema(dataset)
                return self.new(ui_pkg, errors)

            if isinstance(pkg_dict, dict):
                pkgname = pkg_dict.get('dataset').get('uri').split('/')[-1]
            else:
                pkgname = pkg_dict.dataset_uri.split('/')[-1]

            # Delay the redirect - otherwise the Solr data might not yet be up-to-date
            ckanext_helpers.wait_for_solr_to_update()

            self.__redirect_to_dataset_page(pkgname)

        except logic.NotAuthorized:
            abort(401, _('Unauthorized to create package %s') % '')
        except logic.NotFound, e:
            abort(404, _('Dataset not found'))
        except dict_func.DataError:
            abort(400, _(u'Integrity Error'))
        except SearchIndexError, e:
            try:
                exc_str = unicode(repr(e.args))
            except Exception:  # We don't like bare excepts.
                exc_str = unicode(str(e))
            abort(500, _(u'Unable to add package to search index.') + exc_str)
        except logic.ValidationError, e:
            errors['fatal'] = e.error_dict
            error_summary = e.error_summary
            return self.new(ui_util.transform_dcat_schema_to_form_schema(context['package']), errors, error_summary)

    def update(self, id, data=None, errors=None, error_summary=None):
        '''

        :param dict id:
        :param DatasetDcatApOp data:
        :param dict errors:
        :param error_summary:
        :return:
        '''
        log.debug('Update dataset, ID: {0}'.format(id))
        locale = tk.request.environ['CKAN_LANG'] or config.get('ckan.locale_default', 'en')
        package_type = self.__get_package_type()

        context = {'model': 'DCATAP', 'session': model.Session,
                   'user': c.user or c.author, 'auth_user_obj': c.userobj,
                   'save': 'save' in request.params,
                   'validate': 'validate' in request.params,
                   'moderated': config.get('moderated'),
                   'pending': True}

        data_dict = {'id': id}
        context['message'] = 'Update dataset {0} by {0}'.format(data_dict['id'], context['user'])
        try:
            tmp_context = context.copy()
            tmp_context['model'] = model
            logic.check_access('package_update', tmp_context, data_dict)
        except logic.NotAuthorized, e:
            abort(401, _('User %r not authorized to edit %s') % (c.user.encode('ascii', 'ignore'), id))

        if context['save'] and not data:
            request.POST.pop('save')
            for key in request.POST.keys():
                if 'save' == key:
                    del request.POST[key]
                    break
            return self._save_update_and_redirect_to_last_search(id, context,
                                                                 package_type=package_type)  # mdt_package_update!!!

        if context['validate']:
            request.POST.pop('validate')
            for key in request.POST.keys():
                if 'validate' == key:
                    del request.POST[key]
                    break

            data = data or self.__transform_to_data_dict(request.POST)
            context['internal'] = True
            pkg_dict = logic.get_action('package_show')(context, {'uri': data.get('uri')})
            old_dataset = context['package']  # type: DatasetDcatApOp
            old_dataset.update_dataset_for_package_dict(data, {}, context)

            old_dataset, errors = validation.validate_dacat_dataset(old_dataset, context)
            dataset_id = old_dataset.dataset_uri.split('/')[-1]

            return self.update(dataset_id, old_dataset, errors)

        try:
            if not data:
                context['internal'] = True
                # context['validate'] = True
                logic.get_action('package_show')(context, {'id': id})
                c.pkg = context.get("package")
                c.pkg_dict = old_data = ui_util.transform_dcat_schema_to_form_schema(c.pkg, locale)
                data = old_data

                # old data is from the database and data is passed from the
                # user if there is a validation error. Use users data if there.
            else:
                c.pkg = data
                c.pkg_dict = data = ui_util.transform_dcat_schema_to_form_schema(data)


        except logic.NotAuthorized:
            abort(401, _('Unauthorized to read package %s') % '')
        except logic.NotFound:
            abort(404, _('Dataset not found'))

        language = tk.request.environ['CKAN_LANG'] or config.get('ckan.locale_default', 'en')

        data['resources_types'] = self.RESOURCES_TYPES
        c.resources_types_documentation = list(filter(lambda x: x[0] != "http://publications.europa.eu/resource/authority/documentation-type/VISUALIZATION",
                                                      retrieve_all_documentation_types(language)))
        c.resources_types_distribution = list(filter(lambda x: x[0] != "http://publications.europa.eu/resource/authority/distribution-type/VISUALIZATION",
                                                      retrieve_all_distribution_types(language)))
        c.resources_types_visualization = self.RESOURCES_TYPES_VISUALIZATION

        catalog_list = get_action('catalogue_list')(context, {})
        language = tk.request.environ['CKAN_LANG'] or config.get('ckan.locale_default', 'en')

        catalogs = {}
        for key, value in catalog_list.items():
            title = next((title.value_or_uri for title in value.schema.title_dcterms.values() if title.lang == language), key)
            catalogs[key] = title

        c.catalogs = catalogs

        # are we doing a multiphase add?
        if data.get('state', '').startswith('draft'):
            c.form_action = core_helpers.url_for(controller='package', action='new')
            c.form_style = 'new'
            return self.new(data=data, errors=errors, error_summary=error_summary)

        # c.resources_json = core_helpers.json.dumps(data.get('resources', []))

        # convert tags if not supplied in data
        # NOT NEEDED
        # if data and not data.get('tag_string'):
        #     data['tag_string'] = ', '.join(core_helpers.dict_list_reduce(
        #         c.pkg_dict.get('tags', {}), 'name'))
        errors = errors or {}
        vars = {'data': data, 'errors': errors,
                'error_summary': error_summary, 'action': 'edit'}
        # c.errors_json = core_helpers.json.dumps(errors)

        self._setup_template_variables(context, {'id': id}, package_type=package_type)
        # c.related_count = c.pkg.related_count

        # we have already completed stage 1
        # vars['stage'] = ['active']
        # if data.get('state') == 'draft':
        #     vars['stage'] = ['active', 'complete']
        # elif data.get('state') == 'draft-complete':
        #     vars['stage'] = ['active', 'complete', 'complete']

        c.form = base.render(self._package_form(package_type=package_type), extra_vars=vars)

        if context['save'] and not errors:
            return base.render('package/read.html')
        return base.render('package/rdft_edit.html')

    def bulk_update(self, ids=None, errors=None, error_summary=None):
        log.debug('Bulk Update datasets')

        package_type = 'bulk_edit'

        context = {'model': model, 'session': model.Session,
                   'user': c.user or c.author, 'auth_user_obj': c.userobj,
                   'save': 'save' in request.params,
                   'validate': 'validate' in request.params,
                   'moderated': config.get('moderated'),
                   'pending': True}

        id_string = request.params.get('ids')
        ids = ids or id_string.split(',')

        if context['save']:
            request.POST.pop('ids')

            data_dict = self.__transform_to_data_dict(request.POST)

            datasets_list = self.retrieve_datasets_for_ids(context, ids)

            patch_and_save_datasets(context, data_dict, datasets_list, request)

            # --------------------here, the pached dataset is already safed---------------------

            field_schema = [key for key in lookup_package_plugin(package_type).show_package_schema().keys() if
                            not re.match('^_', key)]

            field_schema.append('description')
            field_schema.append('doi')
            field_schema.append('private')
            field_schema.append('keyword_string')

            work_groups = ingestion.get_groups_from_databse_by_id(data_dict.get('groups', []))

            for id in ids:
                upd_work_dict = data_dict.copy()
                sub_context = {'model': model, 'session': model.Session,
                               'user': c.user or c.author, 'auth_user_obj': c.userobj, 'use_cache': False}
                sub_context['for_view'] = True
                log.info('Retrieve %s for update' % (id))
                dataset = logic.get_action('package_show')(sub_context, {'id': id})
                dataset['id'] = id

                concepts_eurovoc = upd_work_dict.get('concepts_eurovoc', [])

                if isinstance(concepts_eurovoc, basestring) and '' == concepts_eurovoc:
                    upd_work_dict.pop('concepts_eurovoc')
                    dataset.pop('concepts_eurovoc')
                    concepts_eurovoc = []
                elif isinstance(concepts_eurovoc, basestring):
                    concepts_eurovoc = [concepts_eurovoc]
                # remove empty strings
                concepts_eurovoc = filter(None, concepts_eurovoc)

                if concepts_eurovoc:
                    upd_work_dict['concepts_eurovoc'] = concepts_eurovoc

                old_groups = dataset.get('groups', [])
                # save old groups and eurovoc_domains
                old_eurovoc_domains = [eurovoc for eurovoc in old_groups if 'eurovoc_domain' == eurovoc.get('type', '')]
                [old_groups.remove(eurovoc) for eurovoc in old_eurovoc_domains]

                # transform update values to groups
                upd_groups = upd_work_dict.get('groups', [])
                groups = copy.deepcopy(work_groups)

                if 'domains_eurovoc' in upd_work_dict.keys():
                    domains = ingestion.get_groups_from_databse_by_id(upd_work_dict.pop('domains_eurovoc'))
                    # eurovoc_domains gets updated, apply old groups
                    if not groups:
                        groups += old_groups

                    for domain in domains:
                        groups.append(domain)
                elif groups:
                    # groups get updated, apply old eurovoc_domains
                    groups += old_eurovoc_domains

                if not groups and '' == data_dict.get('groups', ''):
                    dataset.pop('groups', None)
                    upd_work_dict.pop('groups', None)
                else:
                    upd_work_dict['groups'] = [{'id': group.get('id')} for group in groups]

                if upd_work_dict.get('keyword_string'):
                    dataset.pop('keywords', None)

                dataset = ckanext_converters.convert_dict_to_reusable_dict(dataset)
                dataset.update(upd_work_dict)

                dataset_copy = dict(dataset)
                for key, value in dataset_copy.iteritems():
                    if key not in field_schema:
                        dataset.pop(key, None)

                # dataset.pop('keywords', None)
                # dataset = self._assemble_groups_and_eurovocdomains(dataset)
                sub_context['ignore_auth'] = True
                try:

                    logic.check_access('package_update', context, dataset)
                    log.info('Send %s to package_update' % (id))
                    # new_dataset = get_action('package_update')(sub_context, dataset)

                    log.info('Activate the latest revision of %s' % (dataset['id']))
                    # logic.get_action('make_latest_pending_package_active')(sub_context, dataset)
                    # indexer.index_package(dataset, defer_commit=True)
                except logic.NotAuthorized as e:
                    abort(401, _('Unauthorized to read package %s') % id)
                except logic.NotFound, e:
                    abort(404, _('Dataset not found'))
                except dict_func.DataError:
                    abort(400, _(u'Integrity Error'))
                except SearchIndexError, e:
                    try:
                        exc_str = unicode(repr(e.args))
                    except Exception:  # We don't like bare excepts
                        exc_str = unicode(str(e))
                    abort(500, _(u'Unable to update search index.') + exc_str)
                except logic.ValidationError, e:
                    log.error('%s : %r' % (e.error_summary, e.error_dict))
                    if 'error_dict' in e:
                        errors['error'] = e.error_dict
                    if 'error_summary' in e:
                        error_summary = e.error_summary
                    return self.bulk_update(ids, errors, error_summary)

            filter_str = ''
            for work_id in ids:
                filter_str += 'name:{0} '.format(work_id)
            url = core_helpers.url_for(controller='ckanext.ecportal.controllers.user:ECPortalUserController',
                                       action='dashboard')
            redirect_url = ckanext_helpers.url_with_params(url, [('q', filter_str), ('ext_boolean', 'any')])
            ckanext_helpers.wait_for_solr_to_update()
            return base.redirect(redirect_url)

        if context['validate']:
            request.POST.pop('validate')
            request.POST.pop('ids')

            data_dict = self.__transform_to_data_dict(request.POST)

            empty_field_list = json.loads(data_dict.pop('empty_fields', '[]'))

            for field in empty_field_list:
                if not data_dict.get(field):
                    data_dict.pop(field, None)

            errors = self.__validate(context, data_dict, True)

        config['validate'] = True

        # Bulk edit
        is_bulk_edit = not context['save'] or not context['validate']
        if is_bulk_edit:
            datasets_list = self.retrieve_datasets_for_ids(context, ids)
            form_schema_list = retrieve_form_schema_list_for_datasets(datasets_list)

            datasets_keys_diff = retrieve_differences_for_form_schema_list(form_schema_list)

            bulk_filtered_form_schema = retrieve_bulk_edit_form_schema(datasets_keys_diff, form_schema_list)

            data = bulk_filtered_form_schema
            data['ids'] = id_string

            # convert tags if not supplied in data
            errors = errors or {}
            vars = {'data': data, 'errors': errors,
                    'error_summary': error_summary, 'action': 'edit'}
            c.errors_json = core_helpers.json.dumps(errors)

            self._setup_template_variables(context, data, package_type=package_type)
            # c.related_count = c.pkg.related_count

            catalog_list = get_action('catalogue_list')(context, {})
            language = tk.request.environ['CKAN_LANG'] or config.get('ckan.locale_default', 'en')

            catalogs = {}
            for key, value in catalog_list.items():
                title = next((title.value_or_uri for title in value.schema.title_dcterms.values() if title.lang == language), key)
                catalogs[key] = title

            c.catalogs = catalogs

            # we have already completed stage 1
            vars['stage'] = ['active']
            if data.get('state') == 'draft':
                vars['stage'] = ['active', 'complete']
            elif data.get('state') == 'draft-complete':
                vars['stage'] = ['active', 'complete', 'complete']

            c.form = base.render(self._package_form(package_type=package_type), extra_vars=vars)

            return base.render('package/rdft_new.html')

    def retrieve_datasets_for_ids(self, context, ids):
        datasets_list = []
        for id in ids:
            self.check_rdft_update_package(context, id)
            uri = DATASET_URI_PREFIX + id
            dataset = DatasetDcatApOp(uri)
            get_action('package_show')(context, {'id': id})
            dataset = context['package']
            datasets_list.append(dataset)
        return datasets_list

    def check_rdft_update_package(self, context, id):
        try:
            logic.check_access('rdft_package_update', context, {'id': id})
        except logic.NotAuthorized, e:
            abort(401, _('User %r not authorized to bulk edit %s') % (c.user.encode('ascii', 'ignore')),
                  id)

    def _prepare_dict_for_comparison(self, item_dict, key):
        '''For some dicts remove timestamp to compare'''
        if key in 'keywords, concepts_eurovoc':
            item_dict.pop('revision_timestamp', None)

    def _save_update_and_redirect_to_last_search(self, name_or_id, context, package_type=None):

        log.debug('Package save request name: %s POST: %r', name_or_id, request.POST)
        dataset = None
        try:
            data_dict = self.__transform_to_data_dict(request.POST)
            old_data_dict = data_dict.copy()

            context['ignore_auth'] = True
            try:
                pkg_dict_result = logic.get_action('package_update')(context, data_dict)
                dataset = context['package']
                if data_dict.get('doi'):
                    if pkg_dict_result.privacy_state == 'public':
                        get_action('publish_doi')(context, pkg_dict_result)
            except ValidationError as e:
                errors = e.error_dict
                dataset = context['package']
                return self.update(name_or_id, dataset, errors)

            # logic.get_action('make_latest_pending_package_active')(context, pgk_data_dict)
            indexer.index_package(pkg_dict_result)

            c.pkg = dataset
            c.pkg_dict = pkg_dict_result

            # Delay the redirect - otherwise the Solr data might not yet be up-to-date
            # disabled for new rdft and redirekt to dataset page
            # time.sleep(1)

            # Redirect to the last executed search
            # search_params = USER_SEARCH_PARAMS.get(context['user'], None)
            self.__redirect_to_dataset_page(data_dict.get('id'))

        except logic.NotAuthorized:
            abort(401, _('Unauthorized to read package %s') % id)
        except logic.NotFound, e:
            abort(404, _('Dataset not found'))
        except dict_func.DataError:
            abort(400, _(u'Integrity Error'))
        except SearchIndexError, e:
            try:
                exc_str = unicode(repr(e.args))
            except Exception:  # We don't like bare excepts
                exc_str = unicode(str(e))
            abort(500, _(u'Unable to update search index.') + exc_str)
        except logic.ValidationError, e:
            log.error('%s : %r' % (e.error_summary, e.error_dict))
            errors['fatal'] = e.error_dict
            error_summary = e.error_summary
            return self.update(name_or_id, dataset, errors, error_summary)

    def delete(self):
        log.debug("entering DELETE request")

        # Constants definition
        ACTION_CANCEL = 'cancel'
        ACTION_DELETE = 'delete'
        ACTION_CONFIRMATION = 'confirmation'
        redirection_url = json.dumps({'response_type': 'url', 'body': self._get_dashboard_url()})
        context = self._get_basic_context()

        data = self.__transform_to_data_dict(request.POST)
        action = self._get_value_or_redirect(data, 'action')

        if action == ACTION_CANCEL:
            log.debug("DELETE: Cancelling request & redirection to the dashboard")
            # Returning only the URL, the POST success will do the redirection
            return redirection_url

        selected_datasets = self._get_value_or_redirect(data, 'selected_datasets')

        if action == ACTION_CONFIRMATION:
            log.debug(
                "[Delete datasets] [Preparation of list of datasets] [Start] [Datasets:<{0}>]".format(selected_datasets))
            c.datasets_ids = selected_datasets
            c.datasets = []
            c.deletion_failures = []
            for id in selected_datasets.split(","):
                try:
                    data_dict = {'id': id}
                    selected_dataset = None
                    try:
                        pkg_dict = logic.get_action('package_show')(context, data_dict)
                        selected_dataset = context['package']  # type: DatasetDcatApOp
                    except logic.NotFound:
                        log.debug("[Delete datasets] [Preparation of list of datasets] [failed] [Dataset not found] [URI:<{0}>]".format(id))
                        c.deletion_failures.append({'publisher':'Unknown', 'url':id, 'title':"None"})
                        # abort(404, _('Dataset not found'))
                    if selected_dataset:
                        organization_uri = selected_dataset.schema.publisher_dcterms.get('0', SchemaGeneric('fakepublisher/Unknown')).uri
                        dataset_title = "No title"
                        publisher = organization_uri.split('/')[-1]

                        for titlte_dcterms in selected_dataset.schema.title_dcterms.values():
                            if titlte_dcterms.lang == 'en':
                                dataset_title = titlte_dcterms.value_or_uri

                        dataset_url = selected_dataset.schema.uri or id
                        dataset_dict = {}
                        dataset_dict['publisher'] = publisher
                        dataset_dict['title'] = dataset_title
                        dataset_dict['url'] = dataset_url

                        if selected_dataset.has_doi_identifier():
                            log.debug("[Delete datasets] [Preparation of list of datasets] [Can not delete dataset {0}: contains a DOI]".format(id))
                            c.deletion_failures.append(dataset_dict)
                        else:
                            c.datasets.append(dataset_dict)
                except BaseException as e:
                    c.deletion_failures.append(data_dict)
                    log.error('[Delete datasets] [Preparation of list of datasets] [Can not delete dataset] [URI:<{0}>]'.format(id))
                    log.error(traceback.print_exc(e))

            # if len(c.datasets) == 0:
            #     abort(400, _('Cannot delete dataset(s) with DOI identifiers.'))


            return base.render('package/delete_confirmation.html')

        elif action == ACTION_DELETE:
            log.debug('[Delete datasets] [ACTION DELETE] [START] [URIS:<{0}>'.format(selected_datasets))
            for id in selected_datasets.split(","):
                tmp_context = context.copy()
                log.debug("[Delete datasets] [Deleting dataset] [URI:<{0}>]".format(id))
                data_dict = {'id': id}
                try:
                    logic.check_access('package_delete', tmp_context, {'id': id})
                    tmp_context['ignore_auth'] = True
                except logic.NotAuthorized:
                    abort(401, _('Unauthorized to delete package %s') % '')

                try:
                    pkg_dict = logic.get_action('package_show')(tmp_context, {'id': id})
                    selected_dataset = tmp_context['package']  # type: DatasetDcatApOp
                    if not selected_dataset.has_doi_identifier():
                        logic.get_action('package_delete')(tmp_context, {'id': id})
                except logic.NotFound:
                    log.warn('[Delete datasets] [Deleting dataset] [Dataset not found] [Try to Remove from Solr] [URI:<{0}>]'.format(id))
                    from ckanext.ecportal.lib.search.dcat_index import PackageSearchIndex
                    package_index = PackageSearchIndex()
                    try:
                        fake_dataset = DatasetDcatApOp(id)
                        package_index.remove_dict(fake_dataset)
                        log.error(
                            '[Delete datasets] [Dataset not found] [Remove from Solr] [SUCCESS] [URI:<{0}>]'.format(id))
                    except BaseException as e:
                        log.error('[Delete datasets] [Dataset not found] [Remove from Solr] [Failed] [URI:<{0}>]'.format(id))
                        log.error(traceback.print_exc(e))

                    # abort(404, _('Dataset not found'))
                except BaseException as e:
                    log.error('[Delete datasets] [Deleting dataset] [Dataset not found] [URI:<{0}>]'.format(id))
                    log.error(traceback.print_exc(e))

        else:
            log.debug("Action not recognised, redirection to the dashboard")
            self.__redirect_to_dashboard()

        ckanext_helpers.wait_for_solr_to_update()

        # Returning only the URL, the POST success will do the redirection
        return redirection_url

    def change_privacy_state(self):
        log.debug("entering CHANGE PRIVACY STATE request")

        # Constants definition
        ACTION_CANCEL = 'cancel'
        ACTION_SAVE = 'save'
        ACTION_SELECTION = 'selection'
        redirection_url = json.dumps({'response_type': 'url', 'body': self._get_dashboard_url()})
        context = self._get_basic_context()

        data = self.__transform_to_data_dict(request.POST)
        action = self._get_value_or_redirect(data, 'action')

        if action == ACTION_CANCEL:
            log.debug("CHANGE PRIVACY STATE: Cancelling request & redirection to the dashboard")
            # Returning only the URL, the POST success will do the redirection
            return redirection_url

        selected_datasets = self._get_value_or_redirect(data, 'selected_datasets')

        if action == ACTION_SELECTION:
            log.debug("CHANGE PRIVACY STATE: Display the privacy state selection")
            c.datasets_ids = selected_datasets

            # Check if all datasets are either draft or published
            c.datasets_state = 'mixed'

            are_dataset_all_draft = True
            are_dataset_all_published = True

            for dataset_id in selected_datasets.split(","):
                pkg_dict = logic.get_action('package_show')(context, {'id': dataset_id})
                selected_dataset = context['package']  # type: DatasetDcatApOp
                privacy = selected_dataset.privacy_state
                if privacy == PRIVACY_STATE_PRIVATE:
                    are_dataset_all_published = False
                else:
                    are_dataset_all_draft = False

            if are_dataset_all_published:
                c.datasets_published_state = 'published'
            if are_dataset_all_draft:
                c.datasets_published_state = 'draft'

            return base.render('package/privacy_state_selection.html')

        elif action == ACTION_SAVE:
            privacy_state = self._get_value_or_redirect(data, 'privacy-state')
            if privacy_state not in ("False", "True"):
                abort(417, detail=_('Privacy state must be a boolean'))

            log.debug("CHANGE PRIVACY STATE: Save privacy state [%s] for datasets [%s]" % (
                privacy_state, selected_datasets))

            authorized_datasets = []
            unauthorized_datasets = []
            for ds_id in selected_datasets.split(","):
                data_dict = {'id': ds_id}
                try:
                    logic.check_access('package_update', context, data_dict)
                except logic.NotAuthorized, e:
                    unauthorized_datasets.append(ds_id)
                    continue

                authorized_datasets.append(ds_id)

            context['ignore_auth'] = True

            new_privacy = PRIVACY_STATE_PRIVATE if privacy_state == 'True' else PRIVACY_STATE_PUBLIC

            for ds_id in authorized_datasets:

                dataset_dict = logic.get_action('package_show')(context, {'id': ds_id})
                selected_dataset = context['package']  # type: DatasetDcatApOp
                try:

                    # update only if the privacy is not the same, optimazation of the performance
                    # TODO think about the validation
                    if new_privacy != selected_dataset.privacy_state:
                        selected_dataset.privacy_state = new_privacy
                        # if selected_dataset.privacy_state == DCATAPOP_PRIVATE_DATASET:
                        # selected_dataset.add_draft_to_title()
                        success = selected_dataset.save_to_ts()
                        if success:
                            redis_cache.set_value_in_cache(selected_dataset.dataset_uri, pickle.dumps(selected_dataset), pool=redis_cache.DATASET_POOL)
                            redis_cache.flush_all_from_db(redis_cache.MISC_POOL)
                            search.rebuild(ds_id)

                except logic.ValidationError as ex:
                    log.debug("CHANGE PRIVACY STATE: Validation errors at update of dataset [%s]" % ds_id)
                    abort(422, _('Validation error'))

            context.pop('ignore_auth', None)

        else:
            log.debug("Action not recognised, redirection to the dashboard")
            abort(501, detail="Action not recognised")

        ckanext_helpers.wait_for_solr_to_update()

        return redirection_url

    def assign_doi(self):
        # TODO merge this part of code with delete to avoid redundancy
        # Constants definition
        ACTION_CANCEL = 'cancel'
        ACTION_ASSIGN_DOI = 'assign_doi'
        ACTION_CONFIRMATION = 'confirmation'
        redirection_url = json.dumps({'response_type': 'url', 'body': self._get_dashboard_url()})
        context = self._get_basic_context()

        data = self.__transform_to_data_dict(request.POST)
        action = self._get_value_or_redirect(data, 'action')

        if action == ACTION_CANCEL:
            log.debug("ASSIGN_DOI: Cancelling request & redirection to the dashboard")
            # Returning only the URL, the POST success will do the redirection
            return redirection_url

        selected_datasets = self._get_value_or_redirect(data, 'selected_datasets')

        if action == ACTION_CONFIRMATION:
            log.debug("ASSIGN_DOI: Asking for confirmation for the datasets [%s]" % str(selected_datasets))
            c.datasets_ids = selected_datasets

            c.datasets = []
            c.selection_failures = []
            for id in selected_datasets.split(","):
                data_dict = {'id': id}

                try:
                    pkg_dict = logic.get_action('package_show')(context, data_dict)
                    selected_dataset = context['package']  # type: DatasetDcatApOp
                except logic.NotFound:
                    abort(404, _('Dataset not found'))

                organization = selected_dataset.schema.publisher_dcterms.get('0').uri
                dataset_title = "No title"

                for titlte_dcterms in selected_dataset.schema.title_dcterms.values():
                    if titlte_dcterms.lang == 'en':
                        dataset_title = titlte_dcterms.value_or_uri

                dataset_url = selected_dataset.schema.uri
                status = ["status/status"]
                dataset = {}

                if organization is not None:
                    dataset['publisher'] = organization.split('/')[-1]

                if dataset.get('publisher') is None:
                    dataset['publisher'] = 'Unknown'

                if status and type(status) is list:
                    dataset['status'] = "status/status"  # todo put the correct value

                dataset['title'] = dataset_title
                dataset['url'] = dataset_url

                if selected_dataset.has_doi_identifier():
                    log.debug("ASSIGN_DOI: Cannot assign a DOI to dataset %s: contains a DOI" % id)
                    c.selection_failures.append(dataset)
                else:
                    c.datasets.append(dataset)

            if len(c.datasets) == 0:
                abort(400, _('Cannot assign a DOI to dataset(s) with DOI identifiers.'))
            return base.render('package/assign_doi_confirmation.html')

        elif action == ACTION_ASSIGN_DOI:

            authorized_datasets = []
            unauthorized_datasets = []
            for ds_id in selected_datasets.split(","):
                data_dict = {'id': ds_id}
                try:
                    logic.check_access('package_update', context, data_dict)
                except logic.NotAuthorized, e:
                    unauthorized_datasets.append(ds_id)
                    continue
                try:
                    context['for_edit'] = True
                    context['ignore_auth'] = True
                    dataset = logic.get_action('package_show')(context, data_dict)
                    selected_dataset = context['package']  # type: DatasetDcatApOp
                    if selected_dataset.has_doi_identifier():
                        unauthorized_datasets.append(ds_id)
                        continue
                    authorized_datasets.append((ds_id, dataset, selected_dataset))
                except Exception as e:
                    log.debug(e.message)
                    unauthorized_datasets.append(ds_id)
                    continue

            for ds_id, dataset_dict, selected_dataset in authorized_datasets:

                organization = selected_dataset.schema.publisher_dcterms.get('0').uri
                uri = selected_dataset.schema.uri

                if organization is None:
                    log.debug("ASSIGN_DOI: Cannot assign a DOI to dataset %s: has no publisher" % ds_id)
                    continue

                if selected_dataset.has_doi_identifier():
                    log.debug("ASSIGN_DOI: Cannot assign a DOI to dataset %s: contains a DOI" % ds_id)
                    continue

                organization = organization.split('/')[-1]
                try:
                    doi = doi_facade.DOIFacade(constants.DOI_CONFIG)
                    doi_str = doi.generate_doi(organization, uri)

                    log.debug("ASSIGN_DOI: Generated " + doi_str + " DOI to dataset " + ds_id)

                    selected_dataset.set_doi(doi_str)
                    success = selected_dataset.save_to_ts()
                    if success:
                        redis_cache.set_value_in_cache(selected_dataset.dataset_uri, pickle.dumps(selected_dataset), pool=redis_cache.DATASET_POOL)
                        redis_cache.flush_all_from_db(redis_cache.MISC_POOL)
                        search.rebuild(ds_id)
                        if selected_dataset.privacy_state == 'public':
                            get_action('publish_doi')(context, selected_dataset)

                except logic.ValidationError as ex:
                    log.debug("ASSIGN_DOI: Validation errors at update of dataset [%s]" % ds_id)
                    abort(422, _('Validation error'))
                except Exception as ex:
                    log.debug("ASSIGN_DOI: Failure on DOI generation for %s" % ds_id)
                    log.debug(ex.message)

            context.pop('ignore_auth', None)
        else:
            log.debug("Action not recognised, redirection to the dashboard")
            self.__redirect_to_dashboard()

        ckanext_helpers.wait_for_solr_to_update()

        # Returning only the URL, the POST success will do the redirection
        return redirection_url

    def _get_editable_dataset(self, context, id):
        data_dict = {'id': id}
        try:
            context['for_edit'] = True
            dataset = logic.get_action('package_show')(context, data_dict)
        except logic.NotFound:
            abort(404, _('Dataset not found'))

        if not dataset.get('tag_string'):
            tags = dataset.get('tags') or dataset.get('keywords') or {}
            tags_reduced = core_helpers.dict_list_reduce(tags, 'name')
            dataset['tag_string'] = ', '.join(tags_reduced)

        dataset.pop('keywords', None)

        return dataset

    def _assemble_groups_and_eurovocdomains(self, dataset):
        if (dataset.get('groups') or dataset.get('domains_eurovoc')):
            groups_ids = dataset.get('groups') \
                if isinstance(dataset.get('groups'), list) \
                else [dataset.get('groups')]
            domains_ids = dataset.get('domains_eurovoc') \
                if isinstance(dataset.get('domains_eurovoc'), list) \
                else [dataset.get('domains_eurovoc')]
            groups_ids = groups_ids + domains_ids
            groups = []
            if isinstance(groups_ids, unicode):
                groups_ids = [groups_ids]
            for group_id in groups_ids:
                group = {"id": group_id}
                groups.append(group)
            dataset['groups'] = groups
        return dataset

    def __validate(self, context, data_dict, update=False):

        if not data_dict:
            data_dict = self.__transform_to_data_dict(request.POST)

        groups = ingestion.get_groups_from_databse_by_id(data_dict.get('groups', []))
        if isinstance(groups, basestring) and '' != groups:
            groups = [groups]
        elif isinstance(groups, basestring) and '' == groups:
            work_groups = []

        if data_dict.get('domains_eurovoc'):
            domains = ingestion.get_groups_from_databse_by_id(data_dict.pop('domains_eurovoc'))
            for domain in domains:
                groups.append(domain)
        data_dict['groups'] = groups
        data_dict = get_action('validate_dataset')(context, data_dict)

        errors = data_dict.pop('errors')

        #       if update:     try new validation
        #            errors['error'] = dict((key, value) for key, value in errors.get('error', {}).iteritems() if key not in 'url, name' and 'unique' not in value)
        #            errors['fatal'] = dict((key, value) for key, value in errors.get('fatal', {}).iteritems() if
        #                              key not in 'url, name' and 'unique' not in value)

        return errors

    def _get_value_or_redirect(self, data, attribute):
        value = data.get(attribute)
        if value is None:
            log.debug("Attribute [%s] not found in the POST data, redirecting to dashboard" % attribute)
            abort(417, detail="Attribute [%s] not found in the POST data" % attribute)
        return value

    def _get_basic_context(self):
        return {'model': model,
                'session': model.Session,
                'user': c.user or c.author,
                'auth_user_obj': c.userobj}

    def _setup_template_variables(self, context, data_dict, package_type=None):
        return lookup_package_plugin(package_type).setup_template_variables(context, data_dict)

    def _package_form(self, package_type=None):
        # log.debug(str(package_type) + " / " +
        #    str(lookup_package_plugin(package_type)) + " / " +
        #    str(lookup_package_plugin(package_type).package_form()))
        if ckanext_helpers.is_metadatatool_plugin_activated():
            if 'bulk_edit' == package_type:
                return 'package/rdft_bulk_package_form.html'
            else:
                return 'package/rdft_new_package_form.html'

        return lookup_package_plugin(package_type).package_form()

    def _get_dashboard_url(self):
        return core_helpers.url_for(controller='ckanext.ecportal.controllers.user:ECPortalUserController',
                                    action='dashboard')

    def __redirect_to_dashboard(self, search_params=None):
        '''
        :param search_params: a list of 2-element tuples containing the URL parameters and values.
            Every tuple represents a parameter value pair: [(p1, val1), (p2, val2), ...]
        :return: Nothing
        '''
        url = self._get_dashboard_url()
        log.debug('redirect search_params: ' + str(search_params))

        if search_params and len(search_params) > 0:
            base.redirect(ckanext_helpers.url_with_params(url, search_params))
        else:
            base.redirect(url)

    def __redirect_to_dataset_page(self, id):
        """
        :param id: id or name of the dataset
        :return: Nothing
        """
        url = core_helpers.url_for(controller='ckanext.ecportal.controllers.package:ECPORTALPackageController',
                                   action='read', id=id)
        base.redirect(url)

    def __get_package_type(self):
        '''
        Return the standard package type
        '''
        return 'dataset'

    def __transform_to_data_dict(self, reqest_post):
        '''
        Transform the POST body to data_dict usable by the actions ('package_update', 'package_create', ...)
        :param the POST body of the request object
        :return: a dict usable by the actions
        '''

        # Initialize params dictionary
        data_dict = logic.parse_params(reqest_post)

        for key in data_dict.keys():
            if 'template' in key:
                data_dict.pop(key, None)

        # Transform keys like 'group__0__key' to a tuple (like resources, extra fields, ...)
        try:
            data_dict = logic.tuplize_dict(data_dict)
        except Exception, e:
            log.error(e.message)

        # Collect all tuplized key groups in one key containing a list of dicts
        data_dict = dict_func.unflatten(data_dict)
        data_dict = logic.clean_dict(data_dict)

        return data_dict

    def create_rdf_from_data_dict(self, context, data_dict):
        format = context.get('format', 'rdf')
        loader = self._content_type_from_extension(format)
        package_type = 'dataset'
        c.pkg_dict = data_dict
        self._setup_template_variables(context, data_dict,
                                       package_type=package_type)

        # package_saver.PackageSaver().render_package(data_dict, context)

        template = self._read_template(package_type)
        if 'rdf' in format:
            template = template[:template.index('.') + 1] + format
        else:
            template = 'package/custom_manifest.xml'

        try:
            return render(template, loader_class=loader)
        except ckan.lib.render.TemplateNotFound:
            msg = _("Viewing {package_type} datasets in {format} format is "
                    "not supported (template file {file} not found).".format(
                package_type=package_type, format=format, file=template))
            abort(404, msg)

    def package_import(self):
        package_type = self.__get_package_type()

        context = {'model': model, 'session': model.Session,
                   'user': c.user or c.author, 'auth_user_obj': c.userobj,
                   'save': 'save' in request.params,
                   'validate': 'validate' in request.params
                   }

        errors = {}
        error_summary = {}

        vars = {'data': {}, 'errors': errors,
                'error_summary': error_summary,
                'action': 'import'}
        errors_json = core_helpers.json.dumps(errors)
        if errors:
            log.info('Saving the dataset produced these errors:')
            log.info(errors_json)

        c.errors_json = errors_json

        self._setup_template_variables(context, {}, package_type=package_type)

        c.form = base.render('package/rdft_package_import_form.html', extra_vars=vars)
        return base.render('package/rdft_new.html')

    def validate_import_dataset(self):
        try:
            data = self.__transform_to_data_dict(request.POST)
            file_path = data['file_path']
            path_storage = config.get('ofs.storage_dir', '')

            if not path_storage:
                raise Exception("'ofs.storage_dir' parameter is not set")

            full_path = path_storage + FIX_PATH + file_path

            if not os.path.exists(full_path):
                raise Exception('File is not existing')

            result = self._choose_import_method(full_path)
            if isinstance(result, list):
                # for dataset in result:
                #     dataset['publisher'] = dataset['owner_org']
                dict_result = {'response_type': 'table',
                               'response': base.render('snippets/confirmation_table.html',
                                                       extra_vars={'datasets': result,
                                                                   'table_id': 'import-dataset-table',
                                                                   'is_selectable': True})}
                return json.dumps(dict_result)
            else:
                dict_result = {'response_type': 'link', 'response_error': result['errors'],
                               'response': core_helpers.url_for(controller='package',
                                                                action='read',
                                                                id=result['name'])}

                return json.dumps(dict_result)

        except Exception as e:
            log.error(traceback.print_exc())
            os.remove(full_path)
            os.rmdir(os.path.dirname(os.path.realpath(full_path)))
            # message = e.message or e.extra_msg or ', '.join(
            #     "%s = %s" % (key, val) for (key, val) in e.error_summary.iteritems())
            message = e.message or e.extra_msg or e.error_summary
            abort(417, '<span>%s</span>' % message)

    def validation_display_message(self, dic):
        display_error = ''
        fatal = dic.get('errors', {}).get('fatal', {})
        if len(fatal) > 0:

            for key, value in fatal.iteritems():

                if key == 'resources':
                    for error in value:
                        if isinstance(error, tuple):
                            error = error[1]
                        for k, v in error.iteritems():
                            if isinstance(v, list):
                                for e in v:
                                    display_error += e + "<br>"
                            else:
                                display_error += v + "<br>"
                else:
                    for error in value:
                        if isinstance(error, dict):
                            for property, val in error.iteritems():
                                if isinstance(val, list):
                                    for le in val:
                                        display_error += le + "<br/>"
                                else:
                                    display_error += val + "<br/>"
                        else:
                            if isinstance(error, list):
                                for le in error:
                                    display_error += le + "<br/>"
                            else:
                                display_error += error + "<br/>"
            # abort(417, "File contains validation errors:: \n %s" % display_error)

        dict_error = dic.get('errors', {}).get('error', {})
        if len(dict_error) > 0 and dic.get('private', False) in ['False', False, 'f', 'false']:

            for key, value in dict_error.iteritems():

                if key == 'resources':
                    for error in value:
                        if isinstance(error, tuple):
                            error = error[1]
                        for k, v in error.iteritems():
                            if isinstance(v, list):
                                for e in v:
                                    display_error += e + "<br>"
                            else:
                                display_error += v + "<br>"
                else:
                    for error in value:
                        if isinstance(error, dict):
                            for property, val in error.iteritems():
                                if isinstance(val, list):
                                    for le in val:
                                        display_error += le + "<br/>"
                                else:
                                    display_error += val + "<br/>"
                        else:
                            if isinstance(error, list):
                                for le in error:
                                    display_error += le + "<br/>"
                            else:
                                display_error += error + "<br/>"
            # abort(417, "File contains validation errors: \n %s" % display_error)
            # if (display_error):
            #     abort(417, "File contains validation errors: \n %s" % display_error)
        return display_error

    def _import_rdf_path(self, path):
        file = open(path, 'r')
        rdf_dict = self._import_rdf_file(file)
        return rdf_dict

    def _import_rdf_file(self, file):

        def get_dataset_uri_from_rdf_content(content):
            try:
                from rdflib import Graph
                g = Graph()
                try:
                    g.parse(data=content)
                except BaseException as e:
                    raise Exception("Invalid RDF file")
                uris = g.triples((None, None, None))
                list_uris = []
                for uri, predicat, s in uris:
                    if str(predicat) == "http://www.w3.org/1999/02/22-rdf-syntax-ns#type" and "http://www.w3.org/ns/dcat#Dataset" == str(s):
                        if str(uri) != "http://www.w3.org/ns/dcat#Dataset":
                            list_uris.append(str(uri))
                            return list_uris
                return list_uris
            except BaseException as e:
                log.error("Can not find the URI of the dataset")
                raise

        content_file = file.read()
        list_uri = get_dataset_uri_from_rdf_content(content_file)
        list_datsets_to_import = {}
        if list_uri:
            # prepare the data dict
            context = {}
            uri_dataset = list_uri[0]
            id_dataset = uri_dataset.split("/")[-1]
            data_dict = {"rdfFile": content_file,
                         "addReplaces":
                             [
                                 {
                                     "objectUri": uri_dataset,
                                     "addReplace":
                                         {
                                             "objectStatus": ""
                                         },
                                 }
                             ]

                         }
            # Todo need to get the result of validation ot any feedaback

            context = {'model': model, 'session': model.Session,
                       'user': c.user or c.author,
                       'auth_user_obj': c.userobj}
            # validate the dataset
            errors = None
            context['model'] = 'DCATAP'
            # todo treat the validation report
            # data_dict['errors'] = {'fatal': {}, 'error': {}}
            try:
                global_report = package_save(context, data_dict)
            except ValidationError as ve:
                display_error = self.validation_display_message(ve.error_dict)
                abort(417, "File contains validation errors:: \n %s" % display_error)
            report_ingestion_dataset = {}
            try:
                report_ingestion_dataset = next(iter(global_report), {})
                dic = report_ingestion_dataset.get(uri_dataset, {})
            except BaseException as e:
                log.error(traceback.print_exc())
                raise e

            display_error = self.validation_display_message(dic)
            if display_error:
                abort(417, "File contains validation errors: \n %s" % display_error)
            rdf_dic = dic
            rdf_dic['errors'] = display_error
            rdf_dic["name"] = id_dataset
            return rdf_dic

    def _import_zip_files(self, full_path):
        files_list = []
        files = ingestion.read_zip_file_content(full_path)
        for file in files.namelist():
            read_file = files.open(file)
            extension = os.path.splitext(file)[1][1:].strip().lower()
            if extension == 'rdf':
                files_list.append(self._import_rdf_file(read_file))
            elif extension == 'json':
                imported_json_dict = self._import_json_file(read_file)
                from ckanext.ecportal.lib.ui_util import convert_exported_json_to_form_schema
                data_dict = convert_exported_json_to_form_schema(imported_json_dict)
                dataset_dict = {"title": data_dict.get('title', ''), "url": data_dict.get("uri", ''), 'publisher': data_dict.get('organization', ''), "status": "status/"}
                files_list.append(dataset_dict)
            elif extension == 'xlsx':
                workbook = load_workbook(filename=BytesIO(read_file.read()))
                dataset_dict = self._import_excel_file(workbook)
                dataset_dict['publisher'] = dataset_dict.get('owner_org', "")
                dataset_dict["url"] = dataset_dict.get("id", "")
                dataset_dict["status"] = "status/"  # todo put the correct value
                files_list.append(dataset_dict)
            else:
                raise Exception('File extension is not recognized')
        return files_list

    def _import_json_file(self, file):
        try:
            return json.loads(file.read())
        except BaseException as e:
            raise Exception("Invalid json file")

    def _import_json_path(self, full_path, file=None):

        if not file:
            file = open(full_path, 'r')
        imported_json_dict = self._import_json_file(file)
        # todo create the dataset from the json dcatapop
        # conversion json_export to data_dict
        from ckanext.ecportal.lib.ui_util import convert_exported_json_to_form_schema
        data_dict = convert_exported_json_to_form_schema(imported_json_dict)

        errors = self.__create_update_dataset_from_import(data_dict, full_path)

        return errors

        # return self._create_update_dataset(data_dict, full_path, JSON_FORMAT)

    def _import_excel_path(self, full_path, workbook_excel=None):
        try:
            workbook = workbook_excel
            if not workbook:
                workbook = load_workbook(full_path)
        except BaseException as e:
            raise Exception("Invalid Excel file")
        excel_dict = self._import_excel_file_dcatapop(workbook)
        from ckanext.ecportal.lib.ui_util import convert_exported_json_to_form_schema
        data_dict = convert_exported_json_to_form_schema(excel_dict)
        errors = self.__create_update_dataset_from_import(data_dict, full_path)
        return errors

    def _import_excel_file_dcatapop(self, workbook):
        """
        import the excel file with the format dcatapop
        :param workbook:
        :return:
        """
        try:
            dataset_json_format = _convert_excel_dcatapop_to_json(workbook)
            return dataset_json_format
        except BaseException as e:
            log.error("Import_excel_file failed")
            return None

    def _import_excel_file(self, workbook):
        context = {'model': model, 'session': model.Session,
                   'user': c.user or c.author, 'for_view': True,
                   'auth_user_obj': c.userobj}

        dict = {}
        translations = []
        worksheet = workbook['dataset']

        # Dataset
        for row in worksheet.rows:
            for cell in row:
                property = row[0].value
                if cell.col_idx == 2:
                    # Check the values
                    if cell.row > 1:  # don't check header
                        if property in ['concepts_eurovoc', 'keyword_string', 'geographical_coverage', 'language']:
                            if cell.value:
                                dict[property] = cell.value.split(', ')
                        elif property in ('groups', 'domains_eurovoc'):
                            # groups and domains_eurovoc go both into dest_property = 'groups'
                            # but with a different key
                            dest_property = 'groups'
                            if cell.value:
                                value_list = cell.value.split(', ')
                                # 'groups' can be already filled (multiple property handling here) or missing
                                dict[dest_property] = dict.get(dest_property, [])
                                keymap = {'groups': 'name', 'domains_eurovoc': 'title'}
                                for value in value_list:
                                    dict[dest_property].append({keymap[property]: value})

                        else:
                            if cell.value:
                                if property in ['license_id', 'interoperability_level']:
                                    dict[property] = [cell.value]
                                elif property == "organization":
                                    # publisher = ckanext_converters._convert_publisher_to_dict(cell.value)
                                    # if publisher and len(publisher)>0:
                                    #    dict[property] = publisher[0]
                                    dict['owner_org'] = cell.value
                                else:
                                    dict[property] = str(cell.value)
                elif cell.col_idx >= 3:
                    # Translations
                    if cell.row > 1:
                        if cell.value:
                            english = row[1].value
                            lang = worksheet.rows[0][cell.col_idx - 1].value.lower()
                            translation = {"term": english, "term_translation": cell.value, "lang_code": lang}
                            translations.append(translation)

        resources = workbook.get_sheet_names()

        if resources and len(resources) > 1:
            for resource in resources[1:]:
                sheet = workbook.get_sheet_by_name(resource)
                self._import_resource_worksheet(sheet, dict)

        if len(translations) > 0:
            context['ignore_auth'] = True
            get_action("term_translation_update_many")(context, {"data": translations})

        return dict

    def _import_resource_worksheet(self, sheet, dict):
        resources = dict.get('resources', [])
        resource = {}
        for row in sheet.rows:
            property = row[0].value
            for cell in row:
                if cell.col_idx == 2:
                    # Check the values
                    if cell.row > 1:  # don't check header
                        if cell.value:
                            resource[property] = cell.value
                elif cell.col_idx > 3:
                    # Translations
                    if cell.row > 1:
                        if cell.value:
                            lang = sheet.rows[0][cell.col_idx - 1].value.lower()
                            resource[property + '-' + lang] = str(cell.value)
        resources.append(resource)
        dict['resources'] = resources

    def __create_update_dataset_from_import(self, data_dict, fileName, dataset_dcatapop=None, format_extension=None):
        """
        Create the dataset based on the dict imported from the json or excel.
        :param dataset:
        :param data_dict:
        :param fileName:
        :param format_extension:
        :return:
        """

        file_name = os.path.basename(os.path.splitext(fileName)[0])
        r_index = file_name.rfind('_')
        if r_index > -1:
            file_name = file_name[:r_index]

        name = data_dict.get('name', None)
        identifier = data_dict.get('identifier', None)
        if not identifier:
            identifier = file_name

        if name:
            identifier = name

        criteria = {"id": identifier}

        data_dict['name'] = criteria['id']
        # package = super(ECPORTALPackageController, self).read(datas['identifier'])
        context = {'model': model, 'session': model.Session,
                   'user': c.user or c.author,
                   'auth_user_obj': c.userobj}
        try:
            package = get_action('package_show')(context, criteria)

        except NotFound:
            package = None
            log.info("Package " + criteria['id'] + " cannot be found. Creation of a new package.")

        # validate the dataset
        errors = None
        context['model'] = 'DCATAP'
        result_import = {"title": data_dict.get('title', ''), "url": data_dict.get("uri", ""), "publisher": data_dict.get("organization", ""), "status": "status/"}
        try:
            dict_error = {}
            if not package:
                check_access('package_create', context, data_dict)
                result = get_action('package_create')(context, data_dict)
            else:
                id_package = package.get('dataset').get('uri').split('/')[-1]
                data_dict['id'] = id_package
                tmp_context = context.copy()
                tmp_context['model'] = model
                check_access('package_update', tmp_context, data_dict)
                # Verification with ODP rules
                # Update
                context['ignore_auth'] = True
                context['id'] = id_package
                result = get_action('package_update')(context, data_dict)
            dict_error['errors'] = context.get("errors", {})
            display_error = self.validation_display_message(dict_error)

            result_import["errors"] = display_error
            result_import["name"] = data_dict['id']


        except ValidationError as ve:
            display_error = self.validation_display_message({'errors': ve.error_dict})
            abort(417, "File contains validation errors:: \n %s" % display_error)

        except Exception as ve:
            raise ve
            # abort(417, "File contains validation errors:: \n %s" % display_error)
        return result_import

    def _create_update_dataset(self, dic, fileName, format_extension=None):
        file_name = os.path.basename(os.path.splitext(fileName)[0])
        r_index = file_name.rfind('_')
        if r_index > -1:
            file_name = file_name[:r_index]

        name = dic.get('name', None)
        identifier = dic.get('identifier', None)
        if not identifier:
            identifier = file_name

        if name:
            identifier = name

        criteria = {"id": identifier}

        dic['name'] = criteria['id']
        # package = super(ECPORTALPackageController, self).read(datas['identifier'])
        context = {'model': model, 'session': model.Session,
                   'user': c.user or c.author, 'for_view': True,
                   'auth_user_obj': c.userobj}
        try:
            package = get_action('package_show')(context, criteria)

        except NotFound:
            package = None
            log.info("Package " + criteria['id'] + " cannot be found. Creation of a new package.")

        if not dic:
            raise Exception("File is empty")

        context['ignore_auth'] = True
        dic = get_action('validate_dataset')(context, dic)

        if len(dic['errors']['fatal']) > 0:
            display_error = ''
            for key, value in dic['errors']['fatal'].iteritems():

                if key == 'resources':
                    for error in value:
                        if isinstance(error, tuple):
                            error = error[1]
                        for k, v in error.iteritems():
                            for e in v:
                                display_error += e + "<br>"
                else:
                    for error in value:
                        if isinstance(error, dict):
                            for property, val in error.iteritems():
                                if isinstance(val, list):
                                    for le in val:
                                        display_error += le + "<br/>"
                                else:
                                    display_error += val + "<br/>"
                        else:
                            if isinstance(error, list):
                                for le in error:
                                    display_error += le + "<br/>"
                            else:
                                display_error += error + "<br/>"
            abort(417, "File contains validation errors:: \n %s" % display_error)

        if len(dic['errors']['error']) > 0 and dic.get('private', False) in ['False', False, 'f', 'false']:
            display_error = ''
            for key, value in dic['errors']['error'].iteritems():

                if key == 'resources':
                    for error in value:
                        if isinstance(error, tuple):
                            error = error[1]
                        for k, v in error.iteritems():
                            for e in v:
                                display_error += e + "<br>"
                else:
                    for error in value:
                        if isinstance(error, dict):
                            for property, val in error.iteritems():
                                if isinstance(val, list):
                                    for le in val:
                                        display_error += le + "<br/>"
                                else:
                                    display_error += val + "<br/>"
                        else:
                            if isinstance(error, list):
                                for le in error:
                                    display_error += le + "<br/>"
                            else:
                                display_error += error + "<br/>"
            abort(417, "File contains validation errors: \n %s" % display_error)

        list_lang_res = []
        if 'resources' in dic:
            list_lang_res = self.pop_translation(dic['resources'])

        if not package:
            check_access('package_create', context, dic)
            # Not Found
            # Create
            result = get_action('package_create')(context, dic)

        else:
            dic['id'] = package['id']
            check_access('rdft_dataset_update', context, dic)
            # Verification with ODP rules

            # Update
            context['id'] = package['id']
            result = get_action('dataset_update')(context, dic)

        if len(list_lang_res) > 0 and format_extension == EXCEL_FORMAT:
            for resource in result['resources']:
                for translation_resource in dic['resources']:
                    has_id = 'id' in resource and 'id' in translation_resource
                    is_same_resource = has_id and resource['id'] == translation_resource['id']
                    if is_same_resource or (
                            resource['name'] == translation_resource['name'] and resource['description'] ==
                            translation_resource['description'] and resource['size'] == translation_resource[
                                'size']):
                        translation_resource['revision_id'] = resource['revision_id']
                        translation_resource['id'] = resource['id']
                        break

            translation_dict = tk.get_action('resource_term_translation_create')(context, {'list_lang': list_lang_res,
                                                                                           'data_dict': dic})

        return dic

    def _choose_import_method(self, path):
        extension = os.path.splitext(path)[1][1:].strip().lower()
        if extension == 'rdf':
            return self._import_rdf_path(path)
        elif extension == 'json':
            return self._import_json_path(path)
        elif extension == 'xlsx':
            return self._import_excel_path(path)
        elif extension == 'zip':
            return self._import_zip_files(path)
        else:
            raise Exception('File extension is not recognized')

    def import_selected_dataset_in_zip(self):
        try:
            data = self.__transform_to_data_dict(request.POST)
            file_path = data['file_path']
            selected_datasets = data['selected_datasets'].split(',')
            path_storage = config.get('ofs.storage_dir', '')
            if path_storage:
                full_path = path_storage + FIX_PATH + file_path
                if os.path.exists(full_path):
                    dict_list = self._create_update_dataset_from_zip(full_path, selected_datasets)

                    dict_result = {}

                    url = core_helpers.url_for(
                        controller='ckanext.ecportal.controllers.user:ECPortalUserController',
                        action='dashboard')
                    result = ''

                    for i, dict in enumerate(dict_list):
                        if i > 0:
                            result += ' '

                        result += 'name:' + dict['name']
                        # Don't populate 'response_error' dict with "empty" errors
                        if dict.get('errors', None) and \
                                (dict['errors'].get('fatal', None) or \
                                 dict['errors'].get('error', None) or \
                                 dict['errors'].get('warning', None)):
                            if not 'response_error' in dict_result:
                                dict_result['response_error'] = {}
                            if not 'errors' in dict_result['response_error']:
                                dict_result['response_error']['errors'] = {}
                            dict_result['response_error']['errors'][dict['name']] = dict['errors']

                    dict_result['response'] = ckanext_helpers.url_with_params(url, [('q', result),
                                                                                    ('ext_boolean', 'any')])

                    return json.dumps(dict_result)

                else:
                    raise Exception('File is not existing')

            else:
                raise Exception("'ofs.storage_dir' parameter is not set")
        except Exception as e:
            os.remove(full_path)
            os.rmdir(os.path.dirname(os.path.realpath(full_path)))
            message = e.message or e.extra_msg or ', '.join(
                "%s = %s" % (key, val) for (key, val) in e.error_summary.iteritems())
            abort(417, "<span>" + message + "</span>")

    def _create_update_dataset_from_zip(self, full_path, selected_datasets):
        dict_list = []
        files = ingestion.read_zip_file_content(full_path)
        i = 0
        for file in files.namelist():
            if str(i) in selected_datasets:
                format_extension = None
                read_file = files.open(file)
                extension = os.path.splitext(file)[1][1:].strip().lower()
                if extension == 'rdf':
                    dict = self._import_rdf_file(read_file)
                    format_extension = RDF_FORMAT
                elif extension == 'json':
                    dict = self._import_json_path(full_path="", file=read_file)
                    format_extension = JSON_FORMAT
                elif extension == 'xlsx':
                    workbook = load_workbook(filename=BytesIO(read_file.read()))
                    dict = self._import_excel_path(full_path="", workbook_excel=workbook)
                    format_extension = EXCEL_FORMAT
                dict_list.append(dict)
            i += 1
        return dict_list

    def export(self):
        context = {'model': model, 'session': model.Session,
                   'user': c.user or c.author, 'auth_user_obj': c.userobj}
        data = self.__transform_to_data_dict(request.POST)
        format = data['format']
        selected_datasets = data['selected_datasets'].split(',')
        if format is not None and selected_datasets is not None:
            # Start Export
            js = self._export_datasets(selected_datasets, format)
            return js

            # FIXME: Re-work:
            #    * zipping is repeated code
            #    * use methods instead if where senseful (1 method per format)
            #    * use private (__) instead of protected (_) methods


    def _export_datasets(self, datasets_list, format):
        """
        export the selected dataset to the chosen format.
        :param datasets_list:
        :param format:
        :return:
        """

        def _generate_exported_file(list_content, full_path, format_file="rdf"):
            """
            :param list_content:
            :param full_path:
            :param format_file:
            :rtype :str
            """
            st = datetime.datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d %H:%M:%S')
            if len(list_content) > 1:
                content_type = 'application/zip'
                # Zip creation
                file_name = 'export_' + format_file + '_' + st + '.zip'
                file_path = full_path + file_name
                zf = zipfile.ZipFile(file_path, mode='w', compression=zipfile.ZIP_DEFLATED)
                for content_file in list_content:
                    content = content_file['content'].decode('utf-8').encode('utf8', 'replace')

                    path_file_in_zip = full_path + content_file['name'] + "." + format_file
                    with open(path_file_in_zip, 'w') as f:
                        f.write(content)  # content_file['content'].encode('utf-8', 'replace')
                        f.flush()
                        f.close()
                    zf.write(path_file_in_zip, basename(path_file_in_zip))
                    os.remove(path_file_in_zip)
                zf.close()
            else:
                content_type_format_mapping = {'rdf': 'application/rdf+xml', 'json': 'application/json',
                                               'xls': 'application/xml'}
                content_type = content_type_format_mapping[format_file]
                file_name = list_content[0]['name'] + '_' + st + '.' + format_file
                # JSON export
                file_path = full_path + file_name
                with open(file_path, 'w') as f:
                    f.write(list_content[0]['content'])
                    f.flush()
                    f.close()
            return file_path, file_name, content_type

        log.info("EXPORT Datasets. Start exporting")
        st = datetime.datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d %H:%M:%S')
        if len(datasets_list) > 0:
            # correct format
            context = {'model': model, 'session': model.Session,
                       'user': c.user or c.author, 'for_view': True,
                       'auth_user_obj': c.userobj}

            package_list = []
            file_path = ''
            for dataset in datasets_list:
                get_action('package_show')(context, {'id': dataset})
                package_list.append(context['package'])

            try:
                path_storage = config.get('ofs.storage_dir', '')
                if path_storage:
                    full_path = path_storage + FIX_PATH

                formats = [value['format_id'] for value in FORMATS]
                for legal_format in formats:
                    if legal_format == format and format == JSON_FORMAT:
                        # JSON export
                        list_content = []
                        for package in package_list:  # type:  DatasetDcatApOp

                            result = package.get_dataset_as_json()
                            # test form_schema conversion
                            # from ckanext.ecportal.lib.ui_util import transform_dcat_schema_to_form_export_json
                            # export = json.loads(result)
                            # f = transform_dcat_schema_to_form_export_json(export)
                            pass
                            list_content.append(
                                {"name": package.schema.ckanName_dcatapop['0'].value_or_uri, "content": result})
                        file_path, file_name, content_type = _generate_exported_file(list_content, full_path, format)

                    elif legal_format == format and format == RDF_FORMAT:
                        # RDF export
                        rdf_list = []
                        for package in package_list:  # type: DatasetDcatApOp

                            result = package.get_dataset_as_rdfxml()
                            rdf_list.append(
                                {"name": package.schema.ckanName_dcatapop['0'].value_or_uri, "content": result})
                        file_path, file_name, content_type = _generate_exported_file(rdf_list, full_path, format)

                    elif legal_format == format and format == EXCEL_FORMAT:

                        if len(package_list) > 1:
                            # Export Zip file
                            content_type = 'application/zip'
                            # Zip creation
                            file_name = 'export_excel_' + st + '.zip'
                            file_path = full_path + file_name
                            zf = zipfile.ZipFile(file_path, mode='w', compression=zipfile.ZIP_DEFLATED)

                            for package in package_list:  # type: DatasetDcatApOp
                                ckanName = package.schema.ckanName_dcatapop['0'].value_or_uri
                                excel_file_path = full_path + ckanName + ".xlsx"
                                excel_file_path = get_dataset_as_excel(package, excel_file_path)

                                # to test the onversion

                                zf.write(excel_file_path, basename(excel_file_path))
                                os.remove(excel_file_path)

                            zf.close()
                        else:
                            # Export Excel file
                            package = package_list[0]
                            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                            ckanName = package.schema.ckanName_dcatapop['0'].value_or_uri
                            file_name = "export_" + ckanName + "_" + st + "_.xlsx"
                            file_path = full_path + file_name
                            file_path = get_dataset_as_excel(package, file_path)

                (mode, ino, dev, nlink, uid, gid, size, atime, mtime, ctime) = os.stat(file_path)
                log.error("Write the file to be exported")
                with open(file_path, 'r') as f:
                    shutil.copyfileobj(f, base.response)
                base.response.headers['Content-Length'] = str(size).encode()
                base.response.headers['Content-type'] = content_type.encode()
                base.response.headers['Content-Disposition'] = str('attachment; filename=' + file_name).encode()
                return
            except Exception as ex:
                log.error(traceback.print_exc())

            finally:
                try:
                    os.remove(file_path)
                except BaseException as e:
                    log.error("Export dataset. Can not remove file <{0}>".format(file_path))

    def generate_doi(self, uri=None):
        """
        Assign a DOI to a dataset
        :param context:
        :return: a new generated doi
        """
        publisher = request.POST.get("publisher")
        uri = request.POST.get("uri")
        try:
            doi = doi_facade.DOIFacade(constants.DOI_CONFIG)
            orgs = ckanext_helpers.get_organization({"id": publisher})
            if orgs:
                if uri is not None:
                    return doi.generate_doi(orgs[0], uri)
                else:
                    return doi.generate_doi(orgs[0])
        except Exception as e:
            log.error("Error while generating doi")
            log.error(traceback.print_exc())

    def get_citation(self, id, style, language='en'):
        context = {'user': c.user or c.author, 'for_view': True,
                   'auth_user_obj': c.userobj}

        data_dict = {'id': id}
        pkg_dict = get_action('package_show')(context, data_dict)
        doi_str = pkg_dict['doi']
        language = request.GET.get("language") or language
        return ui_util.download_citation(doi_str, style, language)


def _get_mapping_excel_to_json(path_mapping=""):
    """
    get the maping between the name of properties of the exported json and he excel file
    :rtype:dict[str,dict[str,str]]

    """
    try:
        path_mapping = CKAN_PATH + "/ckanext-ecportal/ckanext/ecportal/model/mapping/DS_json_to_excel.json"  # todo clean that
        with open(path_mapping) as json_mapping_json_to_excel:
            mapping_json_to_excel = json.load(json_mapping_json_to_excel)  # type: dict[str, str]
            mapping_excel_to_json = {}
        for res in mapping_json_to_excel:
            res_dict = {}
            for prop_json, prop_excel in mapping_json_to_excel[res].iteritems():
                if len(prop_excel) == 0:
                    prop_excel = prop_json
                res_dict[prop_excel] = prop_json
            mapping_excel_to_json[res] = res_dict

        return mapping_excel_to_json
    except BaseException as e:
        log.error("Mapping excel names to json failed")
        return None


def get_dataset_as_excel(dataset, file_path):
    """

    Convert the dataset to excel file. the returned value is the path of the  excel file to to be downloaded by the user

    :param DatasetDcatApOp dataset:
    :param str file_path:
    :rtype : str
    """

    def convert_dict_to_worksheet(dcatapop_dict, worksheet, order_parameters, mapping_excel_to_json):
        """
        convert a first level of the dict to a workshet
        :param dict dcatapop_dict:
        :param Workbook workbook:
        :param str title_sheet:
        :param dict mapping_excel_to_json
        :return:
        """
        # todo manage exception

        size_dict = len(dcatapop_dict)
        if size_dict > 0:
            worksheet.cell(row=1, column=1, value="Property")
            worksheet.cell(row=1, column=2, value='Value')

            column_language = _worksheet_header_translation_creation(worksheet)
            i = 2
            for parameter in order_parameters:
                value_string = ''
                parameter_as_in_json = mapping_excel_to_json.get(parameter, "")

                if (parameter_as_in_json not in ["resources_visualization", "resources_documentation",
                                                 "resources_distribution", 'extras']) and (
                        '-' not in parameter_as_in_json):
                    worksheet.cell(row=i, column=1, value=parameter)
                    if parameter_as_in_json in dcatapop_dict:
                        values_parameter = dcatapop_dict.get(parameter_as_in_json, '')
                        if isinstance(values_parameter, list):
                            value_string = excel_multi_values_separator.join(unicode(v) for v in (values_parameter))
                        else:
                            try:
                                value_string = unicode(values_parameter)
                            except BaseException as e:
                                pass
                    else:
                        pass
                    worksheet.cell(row=i, column=2, value=value_string)
                    # add multi lingual values
                    for translated_parameter in dcatapop_dict:
                        value_string = ''
                        if (parameter_as_in_json + "-") in translated_parameter and parameter_as_in_json:
                            language = translated_parameter.split("-")[-1]
                            column_number = column_language.get(language, -1)
                            if column_number != -1:
                                values_translated_parameter = dcatapop_dict[translated_parameter]
                                if isinstance(values_translated_parameter, list):
                                    value_string = excel_multi_values_separator.join(
                                        unicode(v) for v in (values_translated_parameter))
                                else:
                                    value_string = unicode(values_translated_parameter)
                                worksheet.cell(row=i, column=column_number, value=value_string)
                i += 1

    # use the exported json as pivot model
    dcatapop_dict = json.loads(dataset.get_dataset_as_json())  # type: dict[str,list|str]
    dataset_parameters = config.get('ckan.dataset.field.order', []).split()
    resource_parameters = config.get('ckan.resource.field.order', []).split()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'dataset'

    order_parameters = dataset_parameters
    mapping_excel_to_json = _get_mapping_excel_to_json()

    convert_dict_to_worksheet(dcatapop_dict, worksheet, dataset_parameters, mapping_excel_to_json['dataset'])
    # convert the resources to scheets
    resources = dcatapop_dict.get("resources_visualization", []) + dcatapop_dict.get("resources_distribution", [])
    i = 0
    forbidden_characters = config.get('ckan.excel.forbidden.characters', [])
    duplicated_name = []
    for resource_dict in resources:  # type: dict[str,str]
        i = i + 1
        # name of the sheet
        title_resource = "".join(resource_dict.get("title_dcterms", ""))
        title_scheet = _create_name_of_sheet(title_resource, forbidden_characters, duplicated_name)

        if title_scheet == "Not_Named_Resource":
            title_scheet += "_" + str(i)
        worksheet_resource = workbook.create_sheet(title_scheet)
        convert_dict_to_worksheet(resource_dict, worksheet_resource, resource_parameters,
                                  mapping_excel_to_json['resources'])
    workbook.save(file_path)
    # to test the import
    dict_json = _convert_excel_dcatapop_to_json(workbook)
    return file_path


def _convert_excel_dcatapop_to_json(workbook):
    """
    To convert the whole excel to the dict. the format of the returned dixct is the same as the dataset export json one
    :param Workbook workbook:
    :return:
    """
    try:
        dataset_json_sheet = {}
        dataset_json_sheet = _convert_sheet_to_json_dict(workbook, 'dataset', 'dataset')
        dataset_dict_to_be_imported = dataset_json_sheet
        distribution_visualization_list = workbook.get_sheet_names()
        if distribution_visualization_list and len(distribution_visualization_list) > 1:
            dataset_dict_to_be_imported['resources_distribution'] = []
            dataset_dict_to_be_imported['resources_visualization'] = []
            for sheet_name in distribution_visualization_list[1:]:
                distribution_visualization = workbook.get_sheet_by_name(sheet_name)
                dist_dict = _convert_sheet_to_json_dict(workbook, sheet_name, 'resources')

                if dist_dict.get('type_dcterms', [""])[
                    0] == "http://publications.europa.eu/resource/authority/distribution-type/VISUALIZATION":
                    dataset_dict_to_be_imported.get('resources_visualization', []).append(dist_dict)
                else:
                    dataset_dict_to_be_imported.get('resources_distribution', []).append(dist_dict)

        return dataset_dict_to_be_imported

    except BaseException as e:
        log.error("Convert excel to json")
        pass


def _convert_sheet_to_json_dict(workbook, sheet_name, type_resource='dataset'):
    '''
    To convert one sheet to a dict to be integrated in the global dict of the dataset. The latter will be used tin the
    final stem of import
    Convert the scheet to the dict
    :param Workbook workbook:
    :param str sheet_name:
    :param str type_resource. one of 'dataset', distribution, document
    :rtype: dict
    '''

    try:
        dict = {}
        translations = []
        worksheet = workbook[sheet_name]

        one_value_string_parameter = ["uri",
                                      'issued_dcterms',
                                      "modified_dcterms",
                                      "landingPage_dcat",
                                      "temporal_coverage_from",
                                      "temporal_coverage_to",
                                      "rights_dcterms",
                                      "checksumValue_spdx"
                                      ]
        one_value_as_list_parameter = ['title_dcterms', 'description_dcterms']

        # get the mapping of naming
        mapping_excel_to_json = _get_mapping_excel_to_json().get(type_resource, {})
        # Dataset
        dataset_dict = {}
        for row in worksheet.rows:
            for cell in row:
                parameter_as_in_excel = row[0].value
                if cell.col_idx == 2:
                    # Check the values
                    if cell.row > 1:  # don't check header
                        parameter_as_in_json = mapping_excel_to_json.get(parameter_as_in_excel, "")
                        if cell.value and parameter_as_in_json:
                            # a string string
                            cell_value = unicode(cell.value)
                            if parameter_as_in_json in one_value_string_parameter or "contactPoint_dcat" in parameter_as_in_json:
                                dataset_dict[parameter_as_in_json] = cell_value
                            elif parameter_as_in_json in one_value_as_list_parameter:
                                dataset_dict[parameter_as_in_json] = [cell_value]
                            else:
                                dataset_dict[parameter_as_in_json] = cell_value.split(
                                    unicode(excel_multi_values_separator))
                elif cell.col_idx >= 3:
                    # Translations
                    if cell.row > 1:
                        parameter_as_in_json = mapping_excel_to_json.get(parameter_as_in_excel, "")
                        if cell.value and parameter_as_in_json:
                            english = row[1].value
                            lang = worksheet.rows[0][cell.col_idx - 1].value.lower()
                            # translation = {"term": english, "term_translation": cell.value, "lang_code": lang}
                            translated_parameter_as_in_json = "{0}-{1}".format(parameter_as_in_json, lang)
                            # todo thecase of the keyword_string
                            dataset_dict[translated_parameter_as_in_json] = [cell.value]
        return dataset_dict

    except BaseException as e:
        log.error("Convert sheet to json failed . {0}".format(sheet_name))
        pass


def _import_resource_worksheet(sheet, dict):
    resources = dict.get('resources', [])
    resource = {}
    for row in sheet.rows:
        property = row[0].value
        for cell in row:
            if cell.col_idx == 2:
                # Check the values
                if cell.row > 1:  # don't check header
                    if cell.value:
                        resource[property] = cell.value
            elif cell.col_idx > 3:
                # Translations
                if cell.row > 1:
                    if cell.value:
                        lang = sheet.rows[0][cell.col_idx - 1].value.lower()
                        resource[property + '-' + lang] = str(cell.value)
    resources.append(resource)
    dict['resources'] = resources
    return dataset_dict


def _create_name_of_sheet(title_resource, forbidden_characters, duplicated_name):
    """
    :param str title_resource:
    :param list forbidden_characters:
    :param list duplicated_name:
    :return:
    """
    try:
        if title_resource:
            worksheet_name = title_resource[:15]
            worksheet_name = str(worksheet_name.encode("utf-8"))
        else:
            worksheet_name = 'Not_Named_Resource'
        worksheet_name = worksheet_name.translate(None, forbidden_characters)
        worksheet_name = worksheet_name.replace(' ', '_')
        worksheet_name = worksheet_name.decode('utf-8')

        wn = 1
        extension = '_' + str(wn)

        while worksheet_name + extension in duplicated_name:
            wn += 1
            extension = '_' + str(wn)

        worksheet_name += extension
        duplicated_name.append(worksheet_name)
        return worksheet_name
    except BaseException as e:
        log.error("[Export To Excel]. _create_name_of_sheet failed. ")
        return "Not_Named_Resource"


def _export_excel(package_dict, file_path):
    context = {'model': model, 'session': model.Session,
               'user': c.user or c.author, 'for_view': True,
               'auth_user_obj': c.userobj}

    package_dict = ckanext_converters.convert_dict_to_reusable_dict(package_dict)
    parameters_number = len(package_dict.items())
    if parameters_number > 0:
        dataset_parameters = config.get('ckan.dataset.field.order', []).split()
        resource_parameters = config.get('ckan.resource.field.order', []).split()

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'dataset'

        worksheet.cell(row=1, column=1, value="Property")
        worksheet.cell(row=1, column=2, value='Value')

        _worksheet_header_translation_creation(worksheet)

        i = 2
        for parameter in dataset_parameters:
            if parameter != 'resources':
                worksheet.cell(row=i, column=1, value=parameter)
            if parameter in package_dict:
                if isinstance(package_dict[parameter], list):
                    if parameter == 'resources':
                        # Create resource
                        _write_worksheets(workbook, package_dict['resources'], resource_parameters)
                    else:
                        if parameter in package_dict:
                            join_value = ''
                            for concept in package_dict[parameter]:
                                if parameter not in ['type_of_dataset', 'keyword_string', 'temporal_granularity',
                                                     'language', 'geographical_coverage', 'interoperability_level',
                                                     'concepts_eurovoc']:
                                    if 'name' in concept:
                                        join_value += concept['name']
                                    elif 'title' in concept:
                                        join_value += concept['title']
                                else:
                                    join_value += concept
                                if concept != package_dict[parameter][-1]:
                                    join_value += ', '
                        else:
                            join_value = ", ".join(package_dict[parameter])

                        worksheet.cell(row=i, column=2, value=str(join_value))

                elif isinstance(package_dict[parameter], dict):
                    worksheet.cell(row=i, column=2, value=package_dict[parameter]['name'])
                else:
                    if parameter in ['title', 'alternative_title', 'description']:
                        for locale in locales:
                            translation = get_action('term_translation_show')(context, {'lang_codes': [locale.language],
                                                                                        'terms': [
                                                                                            package_dict[parameter]]})
                            if len(translation) > 0:
                                worksheet.cell(row=i, column=export_excel_lang_column[locale.language],
                                               value=translation[0]['term_translation'])

                    elif parameter == 'metadata_language':
                        language = ckanext_converters.__convert_language([package_dict[parameter]])
                        if language and len(language) > 0:
                            package_dict[parameter] = language[0]
                    worksheet.cell(row=i, column=2, value=package_dict[parameter])

            i += 1
        workbook.save(file_path)
        return workbook


def _worksheet_header_translation_creation(sheet):
    column = 4
    for locale in locales:
        if locale.language != 'en':
            sheet.cell(row=1, column=column, value=locale.language.upper())
            export_excel_lang_column[locale.language] = column
            column += 1
    return export_excel_lang_column


def _write_worksheets(workbook, resources, resource_parameters):
    forbidden_characters = config.get('ckan.excel.forbidden.characters', [])
    duplicated_name = []
    for resource in resources:
        if 'name' in resource and resource['name']:
            worksheet_name = resource['name'][:15]
            worksheet_name = str(worksheet_name.encode("utf-8"))
        else:
            worksheet_name = 'Not_Named_Resource'
        worksheet_name = worksheet_name.translate(None, forbidden_characters)
        worksheet_name = worksheet_name.replace(' ', '_')

        worksheet_name = worksheet_name.decode('utf-8')

        wn = 1
        extension = '_' + str(wn)

        while worksheet_name + extension in duplicated_name:
            wn += 1
            extension = '_' + str(wn)

        worksheet_name += extension
        duplicated_name.append(worksheet_name)

        _write_resource_worksheet(workbook, worksheet_name, resource, resource_parameters)


def _write_resource_worksheet(workbook, worksheet_name, dict, parameters):
    sheet = workbook.create_sheet(worksheet_name)
    sheet.cell(row=1, column=1, value='Property')
    sheet.cell(row=1, column=2, value='Value')

    _worksheet_header_translation_creation(sheet)
    langs_translation = ckanext_helpers.get_langs_for_resource(dict)

    j = 2
    for parameter in parameters:
        if parameter in dict:
            if isinstance(dict[parameter], list):
                join_value = dict[parameter].join()
                sheet.cell(row=j, column=1, value=parameter)
                sheet.write_string(j, 2, join_value)
            else:
                # Not List parameter
                sheet.cell(row=j, column=1, value=parameter)
                sheet.cell(row=j, column=2, value=dict[parameter])
                if parameter == 'name':
                    sheet.cell(row=j, column=3, value='Actual Resource Name')
                for lang in langs_translation:
                    translation = ckanext_helpers.get_translated_field_without_fallback(dict, lang, parameter)

                    if translation:
                        if isinstance(translation, tuple):
                            sheet.cell(row=j, column=export_excel_lang_column[lang[0]], value=translation[0])
                        else:
                            sheet.cell(row=j, column=export_excel_lang_column[lang[0]], value=translation)
            j += 1

    return sheet


def get_package_with_translation(context, dict):
    list_lang = []
    for locale in core_helpers.get_available_locales():
        list_lang.append(locale.language)

    translation_title = get_action('term_translation_show')(context,
                                                            {'lang_codes': list_lang, 'terms': [dict['title']]})
    translation_alt_title = get_action('term_translation_show')(context, {'lang_codes': list_lang,
                                                                          'terms': [dict.get('alternative_title', '')]})
    translation_description = get_action('term_translation_show')(context, {'lang_codes': list_lang,
                                                                            'terms': [dict.get('description', '')]})

    if translation_title:
        for translation in translation_title:
            dict['title-' + translation['lang_code']] = translation['term_translation']
    if translation_alt_title:
        for translation in translation_alt_title:
            dict['alternative_title-' + translation['lang_code']] = translation['term_translation']
    if translation_description:
        for translation in translation_description:
            dict['description-' + translation['lang_code']] = translation['term_translation']
